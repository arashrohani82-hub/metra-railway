import io
import logging
import os
import re
import time
import urllib.parse
from datetime import datetime

import openpyxl
from pypdf import PdfReader
import requests

import fixed_ods_app as base

app = base.app
legacy = base.legacy
logger = logging.getLogger(__name__)

PUBLIC_URL = os.environ.get(
    "ODS_PUBLIC_URL", "https://web-production-c7adbe.up.railway.app"
).strip().rstrip("/")
WEBHOOK_URL = f"{PUBLIC_URL}/webhook/telegram"


def force_telegram_webhook():
    token = getattr(legacy, "BOT_TOKEN", None)
    if not token:
        logger.error("ODS RUNTIME: TELEGRAM_BOT_TOKEN is missing")
        return

    try:
        me = requests.get(
            f"https://api.telegram.org/bot{token}/getMe",
            timeout=15,
        ).json()
        if me.get("ok"):
            result = me.get("result") or {}
            logger.info(
                "ODS RUNTIME TELEGRAM BOT: ok=true username=@%s id=%s",
                result.get("username", ""),
                result.get("id", ""),
            )
        else:
            logger.error(
                "ODS RUNTIME TELEGRAM BOT: ok=false error=%s",
                me.get("description", "unknown Telegram error"),
            )
            return

        payload = {
            "url": WEBHOOK_URL,
            "allowed_updates": ["message", "callback_query"],
            "drop_pending_updates": False,
        }
        secret = getattr(legacy, "WEBHOOK_SECRET", "")
        if secret:
            payload["secret_token"] = secret

        result = {}
        for attempt in range(3):
            response = requests.post(
                f"https://api.telegram.org/bot{token}/setWebhook",
                json=payload,
                timeout=15,
            )
            result = response.json() if response.content else {}
            if result.get("ok"):
                break
            retry_after = int((result.get("parameters") or {}).get("retry_after") or 0)
            if response.status_code != 429 and not retry_after:
                break
            if attempt < 2:
                time.sleep(max(1, min(retry_after, 5)))
        logger.info(
            "ODS RUNTIME SET WEBHOOK: ok=%s url=%s description=%s",
            result.get("ok"),
            WEBHOOK_URL,
            result.get("description", ""),
        )
    except Exception:
        logger.exception("ODS RUNTIME WEBHOOK SETUP FAILED")


def _find_history_by_project_folder(folder_name):
    for _owner_uid, records in legacy.offers_history.items():
        for ref, record in (records or {}).items():
            data = record.get("data") or {}
            folder = record.get("project_folder") or data.get("project_folder") or ""
            if folder == folder_name:
                return ref, record
    return None, None


def _strip_label(value, labels):
    text = str(value or "").strip()
    for label in labels:
        if text.lower().startswith(label.lower()):
            return text[len(label):].strip()
    return text


def recover_project_metadata_from_onedrive(folder_name):
    """Recover invoice metadata by scanning archived ODS Excel or PDF files."""
    try:
        config = legacy.microsoft_email_config()
        token = legacy.graph_access_token()
        sender = config["EMAIL_SENDER"]
        year_match = re.match(r"^P(\d{2})-", str(folder_name or ""), re.I)
        year = f"20{year_match.group(1)}" if year_match else datetime.now().strftime("%Y")
        project_root = f"Metra Structure Inc/Projects/{year}/{folder_name}"
        encoded_sender = urllib.parse.quote(sender, safe="")

        project_items = legacy.list_onedrive_children(token, sender, project_root)
        candidates = []
        for item in project_items:
            name = str(item.get("name") or "").strip()
            if name.lower().endswith((".xlsx", ".xlsm", ".pdf")):
                candidates.append((f"{project_root}/{name}", item))
            if item.get("folder"):
                subfolder = f"{project_root}/{name}"
                try:
                    for child in legacy.list_onedrive_children(token, sender, subfolder):
                        child_name = str(child.get("name") or "").strip()
                        if child_name.lower().endswith((".xlsx", ".xlsm", ".pdf")):
                            candidates.append((f"{subfolder}/{child_name}", child))
                except Exception as exc:
                    logger.debug("Project subfolder scan skipped %s: %s", subfolder, exc)

        if not candidates:
            logger.warning("No Excel or PDF source found for invoice project %s", folder_name)
            return {}

        candidates.sort(
            key=lambda pair: (
                str(pair[1].get("name") or "").upper().startswith("ODS"),
                str(pair[1].get("name") or "").lower().endswith((".xlsx", ".xlsm")),
                str(pair[1].get("lastModifiedDateTime") or ""),
            ),
            reverse=True,
        )

        def clean_placeholder(value):
            text = str(value or "").strip()
            placeholders = {
                "", "à compléter", "à confirmer", "—", "none",
                "client", "non indiqué", "non indique",
            }
            return "" if text.lower() in placeholders else text

        def download(relative_path):
            encoded_path = urllib.parse.quote(relative_path, safe="/")
            response = requests.get(
                f"https://graph.microsoft.com/v1.0/users/{encoded_sender}/drive/"
                f"root:/{encoded_path}:/content",
                headers={"Authorization": f"Bearer {token}"},
                timeout=30,
            )
            if response.status_code != 200:
                raise RuntimeError(f"download failed ({response.status_code})")
            return response.content

        def from_excel(content, filename):
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb["ODS"] if "ODS" in wb.sheetnames else wb[wb.sheetnames[0]]
            identity = str(ws["B7"].value or "").strip()
            civility = ""
            name = identity
            match = re.match(r"^(M\.|Mme|M\./Mme)\s+(.*)$", identity, re.I)
            if match:
                civility = match.group(1)
                name = match.group(2).strip()
            address = _strip_label(ws["B8"].value, ["Adresse :", "Adresse:"])
            phone = _strip_label(
                ws["B9"].value,
                ["Cell. :", "Cell.:", "Téléphone :", "Téléphone:"],
            )
            email = _strip_label(
                ws["B10"].value,
                ["Courriel :", "Courriel:", "Email :", "Email:"],
            )
            description = str(ws["B47"].value or "").strip()
            ods_num = str(ws["B12"].value or filename.rsplit(".", 1)[0]).strip()
            price = ws["E47"].value
            result = {
                "name": clean_placeholder(name),
                "civility": clean_placeholder(civility),
                "addr": clean_placeholder(address),
                "project_address": clean_placeholder(address),
                "phone": clean_placeholder(phone),
                "email": clean_placeholder(email),
                "desc": clean_placeholder(description),
                "service": clean_placeholder(description),
                "odsNum": clean_placeholder(ods_num),
                "price": float(price or 0),
            }
            if description:
                result["service_lines"] = [
                    line.strip().lstrip("•").strip().rstrip(";")
                    for line in description.splitlines() if line.strip()
                ][:5]
            return result

        def from_pdf(content, filename):
            reader = PdfReader(io.BytesIO(content))
            text = "\n".join((page.extract_text() or "") for page in reader.pages)
            lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines() if line.strip()]
            address = phone = email = name = civility = ods_num = ""
            for index, line in enumerate(lines):
                lowered = line.lower()
                if lowered.startswith("adresse"):
                    address = line.split(":", 1)[1].strip() if ":" in line else ""
                elif lowered.startswith(("cell.", "cell :", "téléphone", "telephone")):
                    phone = line.split(":", 1)[1].strip() if ":" in line else ""
                elif lowered.startswith(("courriel", "email")):
                    email = line.split(":", 1)[1].strip() if ":" in line else ""
                elif re.match(r"^ODS\d{2}-\d{3}", line, re.I):
                    ods_num = line.split()[0]
                elif re.match(r"^(M\.|Mme|M\./Mme)\s+", line, re.I):
                    match = re.match(r"^(M\.|Mme|M\./Mme)\s+(.*)$", line, re.I)
                    if match and not name:
                        civility, name = match.group(1), match.group(2).strip()
            if not email:
                match = re.search(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", text, re.I)
                email = match.group(0) if match else ""
            result = {
                "name": clean_placeholder(name),
                "civility": clean_placeholder(civility),
                "addr": clean_placeholder(address),
                "project_address": clean_placeholder(address),
                "phone": clean_placeholder(phone),
                "email": clean_placeholder(email),
                "odsNum": clean_placeholder(ods_num or filename.rsplit(".", 1)[0]),
            }
            return result

        ignored_tokens = {
            "reamenagement", "structural", "structure", "projet", "project",
            "conception", "inspection", "design",
        }
        project_tokens = {
            token for token in re.findall(r"[a-z0-9]{3,}", str(folder_name).lower())
            if token not in ignored_tokens and not re.fullmatch(r"p?\d+", token)
        }
        scored_sources = []
        for relative_path, item in candidates:
            filename = str(item.get("name") or "")
            try:
                content = download(relative_path)
                if filename.lower().endswith((".xlsx", ".xlsm")):
                    recovered = from_excel(content, filename)
                else:
                    recovered = from_pdf(content, filename)
                name_tokens = set(re.findall(
                    r"[a-z0-9]{3,}", str(recovered.get("name") or "").lower()
                ))
                metadata_tokens = set(re.findall(
                    r"[a-z0-9]{3,}",
                    " ".join(str(value or "") for value in recovered.values()).lower(),
                ))
                name_matches = len(project_tokens & name_tokens)
                metadata_matches = len(project_tokens & metadata_tokens)
                completeness = sum(
                    bool(recovered.get(field))
                    for field in ("name", "addr", "phone", "email", "odsNum")
                )
                score = name_matches * 100 + metadata_matches * 10 + completeness
                if filename.upper().startswith("ODS"):
                    score += 3
                scored_sources.append((score, recovered, filename))
                logger.info(
                    "Invoice metadata candidate project=%s file=%s score=%s name=%s",
                    folder_name, filename, score, recovered.get("name") or "",
                )
            except Exception as exc:
                logger.warning("Invoice metadata source skipped %s: %s", relative_path, exc)

        if not scored_sources:
            return {}
        scored_sources.sort(key=lambda entry: entry[0], reverse=True)
        best_score, best, best_filename = scored_sources[0]
        logger.info(
            "Selected invoice source project=%s file=%s score=%s client=%s email=%s",
            folder_name,
            best_filename,
            best_score,
            best.get("name") or "",
            best.get("email") or "",
        )
        return best
    except Exception:
        logger.exception("Unable to recover project metadata from files for %s", folder_name)
        return {}

INVOICE_CLIENT_FIELDS = (
    ("name", "nom complet du client"),
    ("addr", "adresse de facturation complète"),
    ("phone", "numéro de téléphone"),
    ("email", "adresse courriel"),
)


def _missing_invoice_client_field(data):
    for field, label in INVOICE_CLIENT_FIELDS:
        value = str(data.get(field) or "").strip()
        if not value:
            return field, label
        if field == "email" and not legacy.valid_client_email(value):
            return field, label
    return None, None


def _ask_missing_invoice_client_field(chat_id, uid, data):
    field, label = _missing_invoice_client_field(data)
    if not field:
        data.pop("waiting_invoice_client_field", None)
        return False
    data["waiting_invoice_client_field"] = field
    legacy.user_data[str(uid)] = data
    legacy.save_user_data()
    examples = {
        "name": "Exemple : David Fiset",
        "addr": "Exemple : 123 rue Exemple, Montréal (Québec) H1H 1H1",
        "phone": "Exemple : (514) 555-1234",
        "email": "Exemple : client@entreprise.ca",
    }
    legacy.tg(
        chat_id,
        "⚠️ Information manquante pour la facture.\n\n"
        f"Veuillez entrer : {label}\n{examples[field]}",
    )
    return True


def _continue_invoice_after_client_details(chat_id, uid, data):
    if _ask_missing_invoice_client_field(chat_id, uid, data):
        return
    if float(data.get("price") or 0) <= 0:
        data["waiting_invoice_contract_amount"] = True
        legacy.user_data[str(uid)] = data
        legacy.save_user_data()
        legacy.tg(
            chat_id,
            "✅ Coordonnées client complètes.\n\n"
            "💰 Quel est le montant total du contrat avant taxes?\n"
            "Exemple : 5500",
        )
        return
    base.show_invoice_options_multi(chat_id, str(uid))


def onedrive_projects():
    config = legacy.microsoft_email_config()
    token = legacy.graph_access_token()
    sender = config["EMAIL_SENDER"]
    year = datetime.now().strftime("%Y")
    root = f"Metra Structure Inc/Projects/{year}"
    items = legacy.list_onedrive_children(token, sender, root)
    projects = [
        item for item in items
        if item.get("folder") is not None
        and str(item.get("name") or "").upper().startswith(f"P{year[-2:]}-")
    ]
    projects.sort(key=lambda item: str(item.get("name") or ""), reverse=True)
    return projects


def show_onedrive_projects_for_invoice(chat_id, uid):
    uid = str(uid)
    try:
        projects = onedrive_projects()
    except Exception as exc:
        logger.exception("Unable to list OneDrive projects for invoicing")
        legacy.tg(chat_id, f"❌ Impossible de lire les projets OneDrive : {exc}")
        return

    if not projects:
        legacy.tg(chat_id, "Aucun projet n'a été trouvé dans OneDrive / Projects.")
        return

    session = legacy.user_data.get(uid, {})
    choices = {}
    rows = []
    for index, item in enumerate(projects[:30], start=1):
        folder = str(item.get("name") or "").strip()
        choices[str(index)] = {
            "folder": folder,
            "web_url": item.get("webUrl") or "",
        }
        rows.append([{
            "text": folder[:62],
            "callback_data": f"od_invoice_project:{index}",
        }])

    session["invoice_project_choices"] = choices
    session.pop("selected_onedrive_project", None)
    legacy.user_data[uid] = session
    legacy.save_user_data()
    text = "🧾 Facturation\n\nChoisissez le projet à facturer :"
    if len(projects) > 30:
        text += "\n(30 projets les plus récents sont affichés.)"
    legacy.tg(chat_id, text, rows)


def select_onedrive_project(chat_id, uid, choice_id):
    uid = str(uid)
    session = legacy.user_data.get(uid, {})
    choice = (session.get("invoice_project_choices") or {}).get(str(choice_id))

    if not choice:
        selected = session.get("selected_onedrive_project") or {}
        if str(selected.get("choice_id") or "") == str(choice_id):
            if session.get("waiting_invoice_contract_amount"):
                return
            if float(session.get("price") or 0) > 0:
                base.show_invoice_options_multi(chat_id, uid)
            return
        legacy.tg(chat_id, "❌ Projet introuvable. Ouvrez de nouveau Facturation.")
        return

    folder = choice["folder"]
    ref, record = _find_history_by_project_folder(folder)
    recovered = recover_project_metadata_from_onedrive(folder)
    if record:
        data = legacy.history_data_copy(record.get("data") or {})
        for field, value in recovered.items():
            if value in (None, "", 0):
                continue
            if field == "price" and float(data.get("price") or 0) > 0:
                continue
            data[field] = value
        data["project_folder"] = folder
        data["project_web_url"] = choice.get("web_url") or record.get("project_web_url") or ""
        data["project_created"] = True
        data["selected_offer_ref"] = ref
    else:
        data = {
            **recovered,
            "project_folder": folder,
            "project_web_url": choice.get("web_url") or "",
            "project_created": True,
            "price": float(recovered.get("price") or 0),
        }

    data["pending_invoice"] = None
    data["waiting_invoice_percentage"] = False
    data["waiting_invoice_amount"] = False
    data["invoice_project_choices"] = session.get("invoice_project_choices") or {}
    data["selected_onedrive_project"] = {
        "choice_id": str(choice_id),
        "folder": folder,
        "web_url": choice.get("web_url") or "",
    }
    legacy.user_data[uid] = data
    legacy.save_user_data()

    if _ask_missing_invoice_client_field(chat_id, uid, data):
        return

    if float(data.get("price") or 0) <= 0:
        data["waiting_invoice_contract_amount"] = True
        legacy.user_data[uid] = data
        legacy.save_user_data()
        client_note = f"\nClient détecté : {data['name']}"
        legacy.tg(
            chat_id,
            f"✅ Projet sélectionné : {folder}{client_note}\n\n"
            "💰 Quel est le montant total du contrat avant taxes?\n"
            "Exemple : 5500",
        )
        return

    base.show_invoice_options_multi(chat_id, uid)


_original_handle_update = legacy.handle_update


def handle_update_runtime(data):
    try:
        msg = data.get("message", {})
        cb = data.get("callback_query", {})

        if msg:
            actor_id = msg.get("from", {}).get("id")
            chat_id = msg.get("chat", {}).get("id")
            text = msg.get("text", "")

            if text in ("🧾 Facturation", "🧾 Facturer un projet"):
                if actor_id not in legacy.ALLOWED_USERS:
                    if chat_id:
                        legacy.tg(chat_id, "⛔ Ce bot est privé.")
                    return
                show_onedrive_projects_for_invoice(chat_id, str(actor_id))
                return

            uid = str(actor_id) if actor_id is not None else ""
            session = legacy.user_data.get(uid, {})
            if text and session.get("waiting_invoice_client_field"):
                field = session.get("waiting_invoice_client_field")
                value = text.strip()
                if field == "email":
                    value = legacy.valid_client_email(value)
                    if not value:
                        legacy.tg(chat_id, "❌ Adresse courriel invalide. Veuillez la saisir de nouveau.")
                        return
                if not value:
                    legacy.tg(chat_id, "❌ Cette information ne peut pas être vide.")
                    return
                session[field] = value
                if field == "addr":
                    session["project_address"] = value
                session.pop("waiting_invoice_client_field", None)
                legacy.user_data[uid] = session
                legacy.save_user_data()
                _continue_invoice_after_client_details(chat_id, uid, session)
                return
            if text and session.get("waiting_invoice_contract_amount"):
                try:
                    amount = float(
                        text.strip().replace("$", "").replace(" ", "").replace(",", "")
                    )
                    if amount <= 0:
                        raise ValueError
                    session["price"] = amount
                    session["waiting_invoice_contract_amount"] = False
                    legacy.user_data[uid] = session
                    legacy.save_user_data()
                    base.show_invoice_options_multi(chat_id, uid)
                except Exception:
                    legacy.tg(chat_id, "❌ Montant invalide. Exemple : 5500")
                return

        if cb:
            cdata = cb.get("data", "")
            if cdata.startswith("od_invoice_project:"):
                actor_id = cb.get("from", {}).get("id")
                chat_id = cb.get("message", {}).get("chat", {}).get("id")
                if actor_id not in legacy.ALLOWED_USERS:
                    if chat_id:
                        legacy.tg(chat_id, "⛔ Ce bot est privé.")
                    return
                try:
                    requests.post(
                        f"https://api.telegram.org/bot{legacy.BOT_TOKEN}/answerCallbackQuery",
                        json={"callback_query_id": cb.get("id")},
                        timeout=3,
                    )
                except Exception:
                    pass
                select_onedrive_project(
                    chat_id,
                    str(actor_id),
                    cdata.split(":", 1)[1],
                )
                return
    except Exception:
        logger.exception("ODS runtime invoice routing failed")

    return _original_handle_update(data)


legacy.handle_update = handle_update_runtime
logger.info("ODS RUNTIME ONEDRIVE INVOICE PROJECT LIST ACTIVE")

force_telegram_webhook()
