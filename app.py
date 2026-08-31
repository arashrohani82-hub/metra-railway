import urllib.parse
import re
import hmac
import html
import copy
import os, json, io, shutil, base64, logging, threading, time
from datetime import datetime
from flask import Flask, request, jsonify, send_file, Response
import anthropic
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, PageBreak, BaseDocTemplate, PageTemplate, Frame
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
import openpyxl
from openpyxl.workbook.properties import CalcProperties
import random
import requests as req
from technical_content import parse_custom_technical_content
from invoice_engine import generate_invoice_pdf, invoice_filename, invoice_values
from concurrent.futures import ThreadPoolExecutor

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
executor = ThreadPoolExecutor(max_workers=10)
project_creation_lock = threading.Lock()
invoice_creation_lock = threading.Lock()
ods_list_lock = threading.Lock()
photo_batch_lock = threading.Lock()
photo_batches = {}

DISPLAY_BRAND = 'Metra Consultation Inc.'
DISPLAY_BRAND_SHORT = 'Metra Consultation'

ANTHROPIC_KEY = os.environ.get('ANTHROPIC_API_KEY')
BOT_TOKEN = os.environ.get('TELEGRAM_BOT_TOKEN')
WEBHOOK_SECRET = os.environ.get('WEBHOOK_SECRET', '')
SETUP_SECRET = os.environ.get('SETUP_SECRET', '')
# Microsoft application and tenant identifiers are not credentials. Keep known
# values as safe fallbacks in case Railway has a pending/empty variable value.
MS_TENANT_ID = os.environ.get(
    'MS_TENANT_ID', 'ac6988d0-6195-40e0-bdaf-ecc951dbd4fa'
)
MS_CLIENT_ID = os.environ.get(
    'MS_CLIENT_ID', 'cc816a89-9af1-4afc-a08d-1ea9dd604ec5'
)
MS_CLIENT_SECRET = os.environ.get('MS_CLIENT_SECRET', '')
EMAIL_SENDER = os.environ.get('EMAIL_SENDER', 'arash.rohani@metrastructure.ca')
INVOICE_EMAIL_SENDER = os.environ.get(
    'INVOICE_EMAIL_SENDER', 'accounting@metrastructure.ca'
)
INVOICE_CC_EMAIL = os.environ.get(
    'INVOICE_CC_EMAIL', 'arash.rohani@metrastructure.ca'
).strip()
INVOICE_DRIVE_OWNER = os.environ.get(
    'INVOICE_DRIVE_OWNER', EMAIL_SENDER
)
INVOICE_FOLDER = os.environ.get(
    'INVOICE_FOLDER', 'Metra Structure Inc/Financial'
).strip('/')
ODS_LIST_PATH = os.environ.get(
    'ODS_LIST_PATH', 'Metra Structure Inc/Offre de service/List.xlsx'
).strip('/')
ALLOWED_USERS = {
    int(value.strip())
    for value in os.environ.get('ALLOWED_TELEGRAM_USER_IDS', '').split(',')
    if value.strip().isdigit()
}
DATA_DIR = os.environ.get('DATA_DIR', '/data')
os.makedirs(DATA_DIR, exist_ok=True)
client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)

W, H = letter
BLACK = colors.black
BASE = os.path.dirname(os.path.abspath(__file__))

LOGOS = {
    'metra': os.path.join(BASE, 'logo_metra.png'),
    'ing':   os.path.join(BASE, 'logo_ing.png'),
    'peo':   os.path.join(BASE, 'logo_peo.png'),
    'rgcq':  os.path.join(BASE, 'logo_rgcq.png'),
}

PRICES = {
    'Analyse structurale générale': 3500,
    'Inspection et rapport structural': 2800,
    "Avis d'expert — stabilisation et renforcement": 3200,
    'Enlèvement de mur porteur': 3800,
    'Inspection des fondations': 2500,
    'Évaluation des fissures et désordres structuraux': 2200,
    'Mur de soutènement': 4200,
    'Conception structurale complète': 6500,
    'Analyse structurale — sous-sol et ajout au-dessus du garage': 4500,
    'Réaménagement intérieur avec modification structurale': 3500,
}

user_data = {}
offers_history = {}

# ── Persistent storage (survives Railway restarts) ──
DATA_FILE = os.path.join(DATA_DIR, 'offre_user_data.json')
OFFERS_FILE = os.path.join(DATA_DIR, 'offre_history.json')

def load_user_data():
    global user_data
    try:
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, 'r') as f:
                user_data = json.load(f)
            logger.info(f"Loaded {len(user_data)} sessions from disk")
    except Exception as e:
        logger.warning(f"load_user_data error: {e}")
        user_data = {}

def save_user_data():
    try:
        with open(DATA_FILE, 'w') as f:
            json.dump(user_data, f, ensure_ascii=False)
    except Exception as e:
        logger.warning(f"save_user_data error: {e}")

def load_offers_history():
    global offers_history
    try:
        if os.path.exists(OFFERS_FILE):
            with open(OFFERS_FILE, 'r') as f:
                offers_history = json.load(f)
            logger.info(f"Loaded offer history for {len(offers_history)} users")
    except Exception as e:
        logger.warning(f"load_offers_history error: {e}")
        offers_history = {}

def save_offers_history():
    try:
        with open(OFFERS_FILE, 'w') as f:
            json.dump(offers_history, f, ensure_ascii=False)
    except Exception as e:
        logger.warning(f"save_offers_history error: {e}")

load_user_data()
load_offers_history()

def get_next_project_num():
    try:
        counter_file = os.path.join(DATA_DIR, 'ods_counter.json')
        if os.path.exists(counter_file):
            with open(counter_file, 'r') as f:
                cdata = json.load(f)
            num = cdata.get('counter', 80)
        else:
            num = 80
        next_num = num + 1
        with open(counter_file, 'w') as f:
            json.dump({'counter': next_num}, f)
        return str(next_num).zfill(3)
    except Exception as e:
        logger.error('counter error: ' + str(e))
        return str(random.randint(80, 999))


def build_ods_num(d):
    import datetime as _dt
    yr = _dt.datetime.now().strftime('%y')
    num = str(d.get('project_num') or '000').zfill(3)
    addr = (d.get('addr') or '')
    addr_lines = [l.strip() for l in addr.split('\n') if l.strip()]
    city_line = addr_lines[1] if len(addr_lines) >= 2 else (addr_lines[0] if addr_lines else '')
    A = city_line[0].upper() if city_line else 'X'
    name = (d.get('name') or '')
    B = name[0].upper() if name else 'X'
    service = (d.get('service') or '')
    C = service[0].upper() if service else 'X'
    code = re.sub(r'[^A-Z]', '', str(d.get('file_code') or '').upper())[:3]
    if len(code) != 3:
        code = f"{A}{B}{C}"
    return f"ODS{yr}-{num}-{code}"

def build_short_title(d):
    project_title = (d.get('project_title') or '').strip()
    if project_title:
        cleaned = re.sub(r'[^\wÀ-ÿ-]+', '-', project_title, flags=re.UNICODE)
        return cleaned.strip('-')[:70]
    service = (d.get('service') or '').strip().lower()
    if 'avis' in service:
        return 'Avis-expert'
    if 'inspection des fond' in service:
        return 'Fondations'
    if 'inspection' in service:
        return 'Inspection'
    if 'enl' in service or 'mur porteur' in service:
        return 'Mur-porteur'
    if 'fissure' in service or 'valuation' in service:
        return 'Fissures'
    if 'soutenement' in service or 'sout' in service:
        return 'Soutenement'
    if 'sous-sol' in service or 'sous sol' in service:
        return 'Sous-sol'
    if 'conception' in service:
        return 'Conception'
    if 'amenagement' in service or 'ménagement' in service:
        return 'Reamenagement'
    if 'analyse' in service:
        return 'Analyse-struct'
    words = (d.get('service') or '').split()
    return '-'.join(words[:2]) if words else 'Mandat'


MISSING_QUESTIONS = {
    'name':  '✏️ Nom du client?',
    'phone': '📞 Numéro de téléphone?',
    'email': '📧 Adresse courriel?',
    'addr':  '📍 Adresse complète du projet?\n(ex: 247 Rue Beaumont\nGranby QC J2G 8S4\nCanada)',
    'project_num': '🔢 Numéro du projet? (ex: 062)',
    'delai': '⏱️ Délai estimé (jours ouvrables)? (ex: 8)',
    'special_note': '📝 Note spéciale à ajouter? Écrivez la note ou tapez « aucune ».',
}

def get_missing_fields(d):
    # The client name is required. Missing contact details use formal placeholders
    # in the preview, Excel and PDF instead of blocking the ODS workflow.
    name = normalize_client_name(d.get('name'))
    d['name'] = name
    d['civility'] = normalize_civility(d.get('civility'))
    return [] if name and name not in ('-', 'None') else ['name']

def show_format_buttons(chat_id, d):
    try:
        price = d.get('price', 0)
        ods = build_ods_num(d) + '-' + build_short_title(d)
        d['odsNum'] = ods
        save_user_data()
    except Exception as e:
        logger.error('show_format_buttons init error: ' + str(e))
        price = d.get('price', 0)
        ods = d.get('odsNum', 'ODS')
    lines = [
        "✅ *Prêt à générer*",
        "",
        "👤 " + client_identity(d),
        "📍 " + client_contact_fields(d)['address'],
        "📞 " + client_contact_fields(d)['phone'],
        "📧 " + client_contact_fields(d)['email'],
        "🏠 " + (d.get('property_type') or '—'),
        "⏱️ " + (d.get('delai') or '—'),
        "🧾 Taxes : " + (d.get('taxes') or '—'),
        "📝 Note : " + (d.get('special_note') or 'Aucune'),
        "🏗️ Accès structural : " + ('Inclus' if d.get('assumption_access') else 'Non inclus'),
        "🏛️ Exclusion architecte/technologue : " + ('Incluse' if d.get('assumption_architect') else 'Non incluse'),
        "",
        "🔧 " + (d.get('project_title') or d.get('service') or '—'),
        "💰 $" + "{:,}".format(price) + " CAD",
        "📄 " + ods,
        "",
        "Tous les éléments sont confirmés. Générer les fichiers?",
    ]
    msg = "\n".join(lines)
    kb = [
        [{'text': '📦 Générer Excel + PDF', 'callback_data': 'both'}],
        [{'text': '✏️ Changer prix', 'callback_data': 'price'}],
        [{'text': '🔄 Nouveau client', 'callback_data': 'nouveau'}],
    ]
    tg(chat_id, msg, kb)




def ask_desc_options(chat_id, uid):
    uid = str(uid)
    d = user_data.get(uid, {})
    try:
        tg(chat_id, "✍️ Préparation du contenu technique de l’ODS...")
        service = d.get('service', '')
        raw_desc = d.get('desc', '')
        property_type = d.get('property_type', '')
        addr = d.get('addr', '')
        prompt = (
            f"You prepare structural-engineering offers of service for {DISPLAY_BRAND} "
            "Write in mature, project-specific structural-engineering French. Never invent a test, deliverable, quantity, "
            "investigation extent, code review, design task, or professional commitment that the "
            "client did not request or that the context does not justify. Avoid promotional wording. "
            "Prepare ONE proposal only, using this exact JSON schema: "
            "{\"project_title\":\"short professional title\","
            "\"file_code\":\"relevant 3-letter uppercase code\","
            "\"short_mandate\":\"one short professional paragraph, maximum 70 words\","
            "\"service_lines\":[\"line 1\",\"line 2\",\"line 3\",\"line 4\",\"line 5 if justified\"]}. "
            "Use 3 to 5 service lines. Each line must be a concrete engineering action or deliverable, "
            "short enough for an ODS table, and end without a period. Every line must be a complete "
            "sentence or complete noun phrase. Never use an ellipsis (... or …), never abbreviate a "
            "service line, and never end with an unfinished connector such as 'incluant', 'notamment', "
            "'concernant' or 'quant à'. Order the services logically: "
            "review of available documents, site work/relevé, structural analysis, design or technical "
            "recommendations, and the requested signed/sealed deliverable. Do not add generic filler such "
            "as client coordination, availability, meetings, communications, or project administration unless "
            "the client specifically requested that service. Distinguish visual inspection from structural "
            "analysis and design. Include 'le cas échéant' only for genuinely conditional work. "
            "Client request/context: " + raw_desc + ". "
            "Initially detected service: " + service + ". Property: " + property_type + ". "
            "Project address: " + addr + ". Return ONLY valid JSON."
        )
        response = client.messages.create(
            model="claude-sonnet-4-6", max_tokens=800,
            messages=[{"role": "user", "content": prompt}]
        )
        result = "".join(b.text for b in response.content if hasattr(b, "text"))
        result = result.replace("```json", "").replace("```", "").strip()
        proposal = json.loads(result)
        service_lines = [
            cleaned + ";"
            for line in proposal.get("service_lines", [])[:5]
            if (cleaned := _clean_service_line(line))
        ]
        short_mandate = str(proposal.get("short_mandate") or raw_desc).strip()
        d["project_title"] = str(proposal.get("project_title") or service).strip()
        d["file_code"] = re.sub(
            r"[^A-Z]", "", str(proposal.get("file_code") or "ODS").upper()
        )[:3].ljust(3, "X")
        d["desc"] = short_mandate
        d["service_lines"] = service_lines
        d["desc_options"] = [short_mandate]
        user_data[uid] = d
        save_user_data()
        provisional_name = (
            f"ODS{datetime.now().strftime('%y')}-XXX-{d['file_code']}-"
            f"{d['project_title']}"
        )
        msg_lines = [
            "📋 Proposition technique",
            "",
            "Titre : " + d["project_title"],
            "Nom du fichier : " + provisional_name,
            "",
            "Mandat court :",
            short_mandate,
            "",
            "Services :",
        ]
        msg_lines.extend("• " + line for line in service_lines)
        tg(chat_id, "\n".join(msg_lines))
        tg(
            chat_id,
            "Confirmer ce contenu avant de compléter les paramètres de l’ODS?",
            [
                [{"text": "✅ Confirmer", "callback_data": "desc_0"}],
                [{"text": "✏️ Modifier le contenu technique", "callback_data": "desc_custom"}],
            ],
        )
    except Exception as e:
        import traceback
        logger.error("ask_desc_options error: " + str(e) + "\n" + traceback.format_exc())
        tg(chat_id, "❌ Erreur description: " + str(e))
        # Don't skip — ask user to write manually
        d["waiting_field"] = "desc"
        d["desc_confirmed"] = False
        user_data[uid] = d
        save_user_data()
        tg(chat_id, "✏️ Écrivez la description du mandat manuellement:")

def ask_next_missing(chat_id, uid):
    d = user_data.get(uid, {})
    missing = get_missing_fields(d)
    if missing:
        field = missing[0]
        d['waiting_field'] = field
        user_data[uid] = d
        save_user_data()
        tg(chat_id, MISSING_QUESTIONS[field] + "\n\n_(Tapez /nouveau pour recommencer)_")
    elif not d.get('addr_confirmed'):
        addr = (d.get('addr') or '').strip()
        if not addr or addr in ('—', 'Non indiqué', 'None'):
            # Completely missing — ask
            d['waiting_field'] = 'addr'
            user_data[uid] = d
            save_user_data()
            tg(chat_id, "📍 Adresse complète du projet?\n(ex: 247 Rue Beaumont\nGranby QC H1A 1A1\nCanada)")
        else:
            # Complete — show confirm buttons (no waiting_field set, buttons handle it)
            addr_display = addr.replace('\n', ', ')
            d['waiting_field'] = None
            user_data[uid] = d
            save_user_data()
            kb = [
                [{'text': '✅ Confirmer', 'callback_data': 'addr_ok'},
                 {'text': '✏️ Modifier', 'callback_data': 'addr_edit'}]
            ]
            tg(chat_id, "📍 Adresse détectée :\n" + addr_display + "\n\nCorrect?", kb)
    elif not d.get('desc_confirmed'):
        executor.submit(ask_desc_options, chat_id, uid)
    elif not d.get('project_num'):
        suggested = get_next_project_num()
        d['waiting_field'] = 'project_num'
        d['suggested_num'] = suggested
        user_data[uid] = d
        save_user_data()
        kb = [[{'text': '✅ ' + suggested, 'callback_data': 'num_ok'}],
              [{'text': '✏️ Autre numéro', 'callback_data': 'num_edit'}]]
        tg(chat_id, "🔢 Numéro suggéré: *" + suggested + "*", kb)
    elif not d.get('price_confirmed'):
        d['waiting_field'] = None
        user_data[uid] = d
        save_user_data()
        suggested_price = int(d.get('price') or 0)
        tg(
            chat_id,
            f"💰 Honoraires proposés : {suggested_price:,} $ CAD\nConfirmer?",
            [
                [{"text": "✅ Confirmer le prix", "callback_data": "price_ok"}],
                [{"text": "✏️ Modifier le prix", "callback_data": "price"}],
            ],
        )
    elif not d.get('delai'):
        d['waiting_field'] = 'delai'
        user_data[uid] = d
        save_user_data()
        tg(chat_id, MISSING_QUESTIONS['delai'])
    elif not d.get('special_note_confirmed'):
        tg(
            chat_id,
            "📝 Une note spéciale à ajouter?",
            [[
                {"text": "Aucune", "callback_data": "note_none"},
                {"text": "Ajouter une note", "callback_data": "note_add"},
            ]],
        )
    elif not d.get('assumption_access_confirmed'):
        tg(
            chat_id,
            "🏗️ Ajouter cette hypothèse à l'offre?\n\n"
            "Accès aux éléments structuraux accessibles "
            "(colonnes, poutres, murs porteurs et fondations).",
            [[
                {"text": "✅ Oui, ajouter", "callback_data": "assumption_access_yes"},
                {"text": "❌ Non", "callback_data": "assumption_access_no"},
            ]],
        )
    elif not d.get('assumption_architect_confirmed'):
        tg(
            chat_id,
            "🏛️ Ajouter cette exclusion à l'offre?\n\n"
            "La vérification, la coordination et l'approbation par l'architecte "
            "ou le technologue ne sont pas incluses.",
            [[
                {"text": "✅ Oui, ajouter", "callback_data": "assumption_architect_yes"},
                {"text": "❌ Non", "callback_data": "assumption_architect_no"},
            ]],
        )
    else:
        show_format_buttons(chat_id, d)


def main_menu():
    return {
        'keyboard': [
            [{'text': '📝 Nouvelle offre'}],
            [{'text': '📁 Convertir une offre en projet'}],
            [{'text': '🧾 Facturation'}],
            [{'text': '📷 Envoyer une photo'}, {'text': '📋 Coller un texte'}],
            [{'text': '❓ Aide'}, {'text': '❌ Annuler'}],
        ],
        'resize_keyboard': True,
        'is_persistent': True,
        'input_field_placeholder': 'Photo ou texte du client…',
    }


def tg(chat_id, text, keyboard=None, reply_markup=None):
    payload = {'chat_id': chat_id, 'text': text}
    if reply_markup:
        payload['reply_markup'] = reply_markup
    elif keyboard:
        payload['reply_markup'] = {'inline_keyboard': keyboard}
    try:
        r2 = req.post(
            f'https://api.telegram.org/bot{BOT_TOKEN}/sendMessage',
            json=payload, timeout=15)
        logger.info(f"tg sent: {r2.status_code} to {chat_id}")
        if r2.status_code != 200:
            logger.error(f"tg response: {r2.text[:200]}")
    except Exception as e:
        logger.error(f"tg error: {e}")

def tg_doc(chat_id, buf, filename, caption):
    try:
        buf.seek(0)
        file_bytes = buf.read()
        logger.info(f"tg_doc: sending {filename}, size={len(file_bytes)} bytes to {chat_id}")
        resp = req.post(
            f'https://api.telegram.org/bot{BOT_TOKEN}/sendDocument',
            data={'chat_id': chat_id, 'caption': caption},
            files={'document': (filename, file_bytes)},
            timeout=60)
        logger.info(f"tg_doc response: {resp.status_code} — {resp.text[:200]}")
    except Exception as e:
        import traceback
        logger.error(f"tg_doc error: {e}")
        logger.error(traceback.format_exc())
        tg(chat_id, f"❌ Erreur envoi fichier: {str(e)}")


def offer_reference(data):
    ods = str(data.get('odsNum') or '')
    match = re.search(r'ODS\d{2}-\d{3}-[A-Z]{3}', ods, re.I)
    if match:
        return match.group(0).upper()
    return safe_archive_filename(ods or 'ODS')


def history_data_copy(data):
    clean = json.loads(json.dumps(data, ensure_ascii=False))
    for key in (
        'email_sending', 'project_creating', 'extracting', 'waiting_field',
        'waiting_price', 'waiting_offer_search',
    ):
        clean.pop(key, None)
    return clean


def record_sent_offer(uid, data):
    uid = str(uid)
    ref = offer_reference(data)
    records = offers_history.setdefault(uid, {})
    existing = records.get(ref, {})
    records[ref] = {
        'data': history_data_copy(data),
        'sent_at': data.get('email_sent_at') or existing.get('sent_at') or datetime.now().isoformat(timespec='seconds'),
        'converted': bool(existing.get('converted') or data.get('project_created')),
        'project_folder': existing.get('project_folder') or data.get('project_folder') or '',
        'project_web_url': existing.get('project_web_url') or data.get('project_web_url') or '',
        'status': 'Accept' if data.get('project_created') else (existing.get('status') or 'In process'),
    }
    save_offers_history()
    return ref


def pending_offer_records(uid, query=''):
    query = query.strip().casefold()
    records = offers_history.get(str(uid), {})
    pending = []
    for ref, record in records.items():
        if record.get('converted') or record.get('status') in {'Refused', 'Closed'}:
            continue
        data = record.get('data') or {}
        searchable = ' '.join([
            ref,
            str(data.get('project_title') or ''),
            str(data.get('name') or ''),
            str(data.get('addr') or ''),
        ]).casefold()
        if query and query not in searchable:
            continue
        pending.append((ref, record))
    return sorted(pending, key=lambda item: item[1].get('sent_at') or '', reverse=True)


def show_pending_offers(chat_id, uid, query=''):
    current = user_data.get(str(uid), {})
    if current.get('email_sent_at'):
        record_sent_offer(uid, current)
    pending = pending_offer_records(uid, query)
    rows = []
    for ref, record in pending[:10]:
        title = str((record.get('data') or {}).get('project_title') or 'Projet').strip()
        label = f"{ref} — {title}"[:60]
        rows.append([{'text': label, 'callback_data': f'offer_pick:{ref}'}])
    rows.append([{'text': '🔎 Rechercher par numéro ODS', 'callback_data': 'offer_search'}])
    if pending:
        text = "📁 Offres en attente de conversion\n\nChoisissez une offre :"
        if len(pending) > 10:
            text += "\n(10 plus récentes affichées; utilisez la recherche pour les autres.)"
    else:
        text = "Aucune offre non convertie trouvée."
        if query:
            text += f"\nRecherche : {query}"
    tg(chat_id, text, rows)


def show_offer_conversion_confirmation(chat_id, uid, ref):
    record = offers_history.get(str(uid), {}).get(ref)
    if not record:
        tg(chat_id, "❌ Offre introuvable. Ouvrez de nouveau la liste.")
        return
    if record.get('converted'):
        message = f"✅ Cette offre est déjà convertie : {record.get('project_folder', '')}"
        if record.get('project_web_url'):
            message += f"\n{record['project_web_url']}"
        tg(chat_id, message)
        return
    data = record.get('data') or {}
    tg(
        chat_id,
        "Confirmer la conversion en projet :\n\n"
        f"ODS : {ref}\n"
        f"Client : {data.get('name') or 'Non indiqué'}\n"
        f"Projet : {data.get('project_title') or data.get('service') or 'Non indiqué'}\n"
        f"Envoyée : {record.get('sent_at') or 'Non indiqué'}",
        [
            [{'text': '✅ Créer le projet', 'callback_data': f'offer_convert:{ref}'}],
            [
                {'text': '❌ Refused', 'callback_data': f'offer_status:Refused:{ref}'},
                {'text': '🔒 Closed', 'callback_data': f'offer_status:Closed:{ref}'},
            ],
            [{'text': '⏸ Hold', 'callback_data': f'offer_status:Hold:{ref}'}],
        ],
    )

def draw_header_footer(canvas, doc):
    canvas.saveState()
    canvas.drawImage(LOGOS['metra'], 1.8*cm, H-2.85*cm, width=3.6*cm, height=1.35*cm, preserveAspectRatio=True, mask='auto')
    canvas.setFillColor(BLACK)
    canvas.setFont('Helvetica-Bold', 10)
    canvas.drawCentredString(W/2, H-1.7*cm, 'Ingénierie des structures / Structural Engineering')
    canvas.setFont('Helvetica', 9)
    canvas.drawCentredString(W/2, H-2.2*cm, '1610-1280 Rue Saint-Jacques')
    canvas.drawCentredString(W/2, H-2.7*cm, 'Montréal – Québec- Canada  H3C 0G1')
    canvas.setFillColor(colors.HexColor('#1155CC'))
    canvas.drawCentredString(W/2, H-3.2*cm, 'info@metrastructure.ca | (438) 867-4131')
    canvas.setFillColor(BLACK)
    canvas.setLineWidth(1)
    canvas.line(1.8*cm, H-3.6*cm, W-1.8*cm, H-3.6*cm)
    canvas.setLineWidth(0.5)
    canvas.line(1.8*cm, 2.0*cm, W-1.8*cm, 2.0*cm)
    canvas.drawImage(LOGOS['ing'],  2*cm,  0.3*cm, width=2*cm,  height=1.4*cm, preserveAspectRatio=True, mask='auto')
    canvas.drawImage(LOGOS['peo'],  8*cm,  0.3*cm, width=3*cm,  height=1.4*cm, preserveAspectRatio=True, mask='auto')
    canvas.drawImage(LOGOS['rgcq'],15*cm,  0.3*cm, width=2*cm,  height=1.4*cm, preserveAspectRatio=True, mask='auto')
    canvas.setFont('Helvetica', 8)
    canvas.drawRightString(W-1.8*cm, 0.6*cm, f'{doc.page} | Page')
    canvas.restoreState()

def normalize_client_name(value):
    """Trim repeated whitespace and remove a civility already included in the name."""
    name = re.sub(r'\s+', ' ', str(value or '')).strip()
    name = re.sub(r'^(?:M\.?|Mme|Mlle|Monsieur|Madame)\s+', '', name, flags=re.IGNORECASE)
    return name


def normalize_civility(value):
    """Return only an approved French civility; stay neutral when uncertain."""
    raw = re.sub(r'\s+', '', str(value or '')).lower()
    if raw in ('m', 'm.', 'monsieur', 'mr'):
        return 'M.'
    if raw in ('mme', 'madame', 'mrs', 'ms'):
        return 'Mme'
    return 'M./Mme'


def client_identity(data):
    name = normalize_client_name(data.get('name'))
    civility = normalize_civility(data.get('civility'))
    return f"{civility} {name or 'À compléter'}"


def client_contact_fields(data):
    addr = ', '.join(
        line.strip() for line in str(data.get('addr') or '').split('\n') if line.strip()
    )
    phone = str(data.get('phone') or '').strip()
    email = str(data.get('email') or '').strip()
    return {
        'address': addr or 'À confirmer',
        'phone': phone or 'À compléter',
        'email': email or 'À compléter',
    }


def _clean_service_line(value):
    """Return a complete service line without truncation or ellipsis."""
    text = re.sub(r'\s+', ' ', str(value or '')).strip(' •;.')
    ellipsis = re.search(r'(?:\.{3,}|…)', text)
    if ellipsis:
        text = text[:ellipsis.start()].rstrip(' ,;:–-')
        if ',' in text:
            prefix, fragment = text.rsplit(',', 1)
            fragment = fragment.strip().lower()
            if re.match(
                r"^(?:incluant|notamment|comprenant|y compris|ainsi que|et\b|ou\b|quant à|relatif|relative|concernant)",
                fragment,
            ):
                text = prefix.rstrip(' ,;:–-')
    return text.strip(' •;.')


def _build_service_desc(data):
    """Up to 5 complete technical lines for the PDF table cell."""
    # Priority: service_lines from extraction
    raw = data.get('service_lines') or []
    if not raw:
        import re as _re
        desc = data.get('desc') or data.get('service') or ''
        raw = [p.strip() for p in _re.split(r'[;.\n]', desc) if p.strip()]
    result = []
    for line in raw[:5]:
        compact_line = _clean_service_line(line)
        if not compact_line:
            continue
        result.append('• ' + compact_line + ';')
    return '<br/>'.join(result) if result else data.get('service', '')


def selected_assumptions(data):
    """Return only the project assumptions selected for this offer."""
    assumptions = [
        "Plans architecturaux fournis avant le début du mandat (si disponible);",
    ]
    if data.get('assumption_access'):
        assumptions.append(
            "Accès aux éléments structuraux accessibles "
            "(colonnes, poutres, murs porteurs, fondations);"
        )
    if data.get('assumption_architect'):
        assumptions.append(
            "Vérification, coordination et approbation par l'architecte ou le "
            "technologue non incluses dans la présente offre."
        )
    return [f"{index}- {text}" for index, text in enumerate(assumptions, 1)]

def generate_pdf(data):
    buf = io.BytesIO()
    doc = BaseDocTemplate(buf, pagesize=letter,
        rightMargin=1.8*cm, leftMargin=1.8*cm,
        topMargin=3.8*cm, bottomMargin=2.15*cm)
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='normal')
    doc.addPageTemplates([PageTemplate(id='all', frames=frame, onPage=draw_header_footer)])

    def s(name, font='Helvetica', size=9.5, leading=11.7, align=TA_LEFT, sb=0, sa=0):
        return ParagraphStyle(name, fontName=font, fontSize=size, leading=leading,
                              textColor=BLACK, alignment=align, spaceBefore=sb, spaceAfter=sa)

    sn=s('n'); sb_=s('b',font='Helvetica-Bold'); sh=s('h',font='Helvetica-Bold',sb=4,sa=1.5)
    sr=s('r',align=TA_RIGHT); sj=s('j',leading=11.4,sb=2,sa=2)
    sc=s('cadre',font='Helvetica-Bold',size=9.5,leading=12,sb=4,sa=3,align=TA_CENTER)
    shb=s('hb',font='Helvetica-Bold',align=TA_CENTER); shn=s('hn',align=TA_CENTER); str_=s('tr',align=TA_RIGHT)

    price = float(data.get('price', 3200))
    pf = f'$ {price:,.2f}'
    story = []

    story.append(Paragraph(f'Date :  {data.get("date", datetime.now().strftime("%Y-%m-%d"))}', sr))
    story.append(Spacer(1, 3))
    contact = client_contact_fields(data)
    story.append(Paragraph(client_identity(data), sn))
    story.append(Paragraph('Adresse : ' + contact['address'], sn))
    story.append(Paragraph('Cell. : ' + contact['phone'], sn))
    story.append(Paragraph('Courriel : ' + contact['email'], sn))
    story.append(Spacer(1, 3))
    story.append(Paragraph(f'<b>{data["odsNum"]}</b>', sn))
    story.append(Paragraph('CADRE CONTRACTUEL – PROPOSITION DE SERVICES | METRA CONSULTATION INC.', sc))
    story.append(Paragraph(
        f"L'équipe de {DISPLAY_BRAND} vous remercie pour votre confiance à l'égard de notre proposition de services. "
        "Nous vous informons que la présente offre, ainsi que les conditions qui l'accompagnent, forment un accord unique et indissociable. "
        "Toute acceptation de cette offre vaut acceptation complète et sans réserve de l'ensemble des modalités qui y sont énoncées. "
        "Aux fins des présentes, le terme « Client » réfère à la personne, physique ou morale, qui confie le mandat et qui demeure responsable du paiement des honoraires afférents.", sj))

    for heading, text in [
        ("1. Description des services",f"{DISPLAY_BRAND} offre ses services d'ingénierie-conseil conformément aux cadres légaux, aux normes en vigueur et aux règles professionnelles applicables, notamment celles de l'Ordre des ingénieurs du Québec (OIQ) et de Professional Engineers Ontario (PEO), pour le périmètre défini au mandat. Les services sont fournis selon une obligation de moyens et non de résultat. La responsabilité de {DISPLAY_BRAND} ne pourra excéder, sous réserve des dispositions légales applicables, le montant des honoraires payés pour le présent mandat."),
        ("2. Versement initial","À défaut d'une entente écrite contraire, un acompte représentant 25 % du montant total de l'offre de services est requis au moment de la signature."),
        ("3. Honoraires et modalités de paiement",f"Les honoraires et frais remboursables sont facturés selon la progression des travaux et sont exigibles dès réception de la facture. Tout montant non réglé dans un délai de trente (30) jours sera assujetti à des intérêts de 1,5 % par mois (19,56 % par année). En cas de non-paiement, {DISPLAY_BRAND} pourra suspendre la prestation des services. Les taxes applicables s'ajoutent aux honoraires."),
        ("4. Gestion des retards et arrêt du projet","En cas de suspension ou d'annulation du projet, le client est responsable du paiement des coûts engagés et des prestations réalisées jusqu'à la date de notification écrite."),
        ("5. Cadre contractuel","Ce document tient lieu d'entente complète entre les parties. Aucun changement ne sera valide à moins d'être formulé par écrit."),
    ]:
        story.append(Paragraph(f'<b>{heading}</b>', sh))
        story.append(Paragraph(text, sj))

    sn2=s('n2',size=8.5,leading=10.2)
    sb2=s('b2',font='Helvetica-Bold',size=8.5,leading=10.2)
    sh2=s('h2',font='Helvetica-Bold',size=8.5,leading=10.2,sb=2,sa=1)
    sj2=s('j2',size=8.5,leading=10.2,sb=1,sa=1)
    shb2=s('hb2',font='Helvetica-Bold',size=8,leading=9.3,align=TA_CENTER)
    shn2=s('hn2',size=8,leading=9.3,align=TA_CENTER)
    str2=s('tr2',size=8,leading=9.3,align=TA_RIGHT)

    story.append(Paragraph('<b>6. Présence sur site et logistique</b>', sh2))
    story.append(Paragraph("Toute requête de déplacement doit être transmise au moins 48 heures avant la date prévue.", sj2))
    story.append(Spacer(1, 2))
    story.append(Paragraph('<b>TAUX HORAIRES</b>', sb2))
    story.append(Spacer(1, 2))
    rt = Table([[Paragraph(r, str2), Paragraph(v, sn2)] for r,v in [
        ('Ingénieur senior :','130 $ /h'),('Ingénieur intermédiaire :','110 $ /h'),
        ('Ingénieur junior :','105 $ /h'),('Technicien :','100 $ /h'),('Dessinateur :','85 $ /h'),
    ]], colWidths=[9*cm, 3*cm])
    rt.setStyle(TableStyle([('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2)]))
    story.append(rt)
    story.append(PageBreak())
    story.append(Paragraph('<b>HONORAIRES – FORFAIT DU PROJET</b>', sb2))
    story.append(Spacer(1, 3))
    hon_data = [
        [Paragraph('<b>Description des services</b>',shb2),Paragraph('<b>Unité</b>',shb2),Paragraph('<b>Quantité</b>',shb2),Paragraph('<b>Coût unitaire</b>',shb2),Paragraph('<b>Coût total</b>',shb2)],
        [Paragraph(_build_service_desc(data),s('sd',size=8,leading=9.3)),Paragraph('Forfait',shn2),Paragraph('1',shn2),Paragraph(pf,shn2),Paragraph(pf,shn2)],
        ['','','',Paragraph('Total des honoraires du projet',str2),Paragraph(f'<b>{pf}</b>',shb2)],
    ]
    ht = Table(hon_data, colWidths=[8.5*cm,1.8*cm,1.8*cm,3.8*cm,2.6*cm])
    ht.setStyle(TableStyle([
        ('BOX',(0,0),(-1,-1),0.5,BLACK),('INNERGRID',(0,0),(-1,-1),0.5,BLACK),
        ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
        ('LEFTPADDING',(0,0),(-1,-1),4),('RIGHTPADDING',(0,0),(-1,-1),4),
        ('VALIGN',(0,0),(-1,-1),'TOP'),('SPAN',(0,2),(2,2)),
    ]))
    story.append(ht)
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        '<b>Taxes : </b>' + (data.get('taxes') or 'En sus'),
        sn2,
    ))
    if data.get('special_note'):
        story.append(Spacer(1, 2))
        story.append(Paragraph(
            '<b>Note spéciale : </b>' + str(data.get('special_note'))[:250],
            sn2,
        ))
    story.append(Spacer(1, 4))
    story.append(Paragraph('<b>AUTRES FRAIS (SI APPLICABLE)</b>', sb2))
    story.append(Spacer(1, 3))
    story.append(Paragraph('Le délai de livraison estimé est de ' + (data.get('delai') or '10') + ' jours ouvrables suivant la visite finale sur site.', sn2))
    story.append(Spacer(1, 3))
    story.append(Paragraph('<b>Cette offre est basée sur les hypothèses suivantes :</b>', sb2))
    for h in selected_assumptions(data):
        story.append(Paragraph(h, sn2))
    story.append(Spacer(1, 4))
    story.append(Paragraph("Cette offre est valable 30 jours. Pour l'accepter, veuillez compléter les sections suivantes.", sn2))
    story.append(Spacer(1, 8))
    sig = Table([
        [Paragraph('<b>Arash Rohani</b> , ing., P.Eng.',sn2),Paragraph('<b>Nom du client:</b>',sn2)],
        [Paragraph('Président-Ingénieur en structure',sn2),''],
        [Paragraph(DISPLAY_BRAND,sn2),Paragraph('<b>Date:</b>',sn2)],
    ], colWidths=[9*cm,8.5*cm])
    sig.setStyle(TableStyle([('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2)]))
    story.append(sig)
    doc.build(story)
    buf.seek(0)
    return buf

def generate_excel(data):
    template = os.path.join(BASE, 'template.xlsx')
    out_path = f"/tmp/{data['odsNum']}.xlsx"
    shutil.copy(template, out_path)
    wb = openpyxl.load_workbook(out_path)
    ws = wb['ODS']
    contact = client_contact_fields(data)
    ws['B7'] = client_identity(data)
    ws['B8'] = 'Adresse : ' + contact['address']
    ws['B9'] = 'Cell. : ' + contact['phone']
    ws['B10'] = 'Courriel : ' + contact['email']
    ws['B12'] = f"{data.get('odsNum','ODS')}-{(data.get('name') or 'client').replace(' ','-')}"
    service_lines = data.get('service_lines') or []
    ws['B47'] = '\n'.join(service_lines[:5]) or data.get('desc', data.get('service',''))
    ws['C47'] = 'Forfait'
    ws['D47'] = 1
    ws['E47'] = float(data['price'])
    ws['F47'] = '=E47*D47'
    ws['F48'] = '=SUM(F47:F47)'
    assumptions = selected_assumptions(data)
    for row, value in zip((54, 55, 56), assumptions + ['', '', '']):
        ws.cell(row=row, column=2).value = value
    wb.save(out_path)
    with open(out_path, 'rb') as f:
        buf = io.BytesIO(f.read())
    os.remove(out_path)
    buf.seek(0)
    return buf

PHOTO_BATCH_DELAY_SECONDS = 2.0
MAX_EXTRACTION_IMAGES = 5


def _photo_extraction_prompt(image_count):
    return f"""Analyze ALL {image_count} attached image(s) together as one client request.
Images may be consecutive email screenshots and/or photos of visible building conditions.
Extract only information visibly present. Do not infer a diagnosis or invent missing facts.
Return ONLY JSON:
{{"client_name":"","client_civility":"M.|Mme|M./Mme","phone":"","email":"","address":"","soumission_ref":"","project_description":"","image_observations":[],"property_type":"","suggested_service":"","suggested_price":0}}

CONTACT RULES:
1. Read the complete email header AND the complete message body across every screenshot.
2. From/sender email is the client's email. The To/recipient account belongs to Metra Consultation; never return an @metrastructure.ca address as the client email.
3. An explicit self-identification in the body, such as "my name is" or "je m'appelle", overrides a shortened/different display name in the From line.
4. Extract every visible client phone number from the message body or signature. Do not leave phone blank when a number is visibly written.
5. Use the full project address stated in the body or subject. Preserve street number, street, city and postal code exactly as visible.
6. Do not use a Cc contact when a From contact is visible.

PROJECT RULES:
1. project_description must be a concise but complete 2–4 sentence summary of the client's actual request, existing condition, requested opinion/report/design, and any explicitly requested tests or deliverables.
2. Review every attached building photo. Put only directly visible conditions in image_observations (for example crack location or affected element); do not claim a cause, severity, movement, structural safety, or required repair from a photo alone.
3. Incorporate relevant image_observations into project_description so the technical scope is based on both the written request and the photos.
4. Never replace a specific request with generic engineering language.

For client_civility: use "M." for a confidently male first name, "Mme" for a confidently female first name, and "M./Mme" when confidence is low.
suggested_service must be one of: "Analyse structurale générale","Inspection et rapport structural","Avis d'expert — stabilisation et renforcement","Enlèvement de mur porteur","Inspection des fondations","Évaluation des fissures et désordres structuraux","Mur de soutènement","Conception structurale complète","Analyse structurale — sous-sol et ajout au-dessus du garage","Réaménagement intérieur avec modification structurale".
suggested_price is only a preliminary internal suggestion in CAD. ONLY JSON."""


def _contact_verification_prompt():
    return """Re-read all attached images only to verify client contact facts.
Return ONLY JSON:
{"client_name":"","client_civility":"M.|Mme|M./Mme","phone":"","email":"","address":""}
Rules:
- An explicit self-identification in the message body (for example "My name is Nevin El-Tahry") overrides a different or shortened From display name.
- From/sender email is the client email. Never use To, Cc, or an @metrastructure.ca address.
- Capture a phone number written anywhere in the body or signature.
- Capture the complete visible project address from the body or subject.
- Do not invent missing text. ONLY JSON."""


def _model_json(response):
    text = ''.join(b.text for b in response.content if hasattr(b, 'text'))
    return json.loads(text.replace('```json', '').replace('```', '').strip())


def _download_telegram_images(file_ids):
    images = []
    for file_id in file_ids[:MAX_EXTRACTION_IMAGES]:
        result = req.get(
            f'https://api.telegram.org/bot{BOT_TOKEN}/getFile?file_id={file_id}',
            timeout=10,
        ).json()
        file_path = result['result']['file_path']
        response = req.get(
            f'https://api.telegram.org/file/bot{BOT_TOKEN}/{file_path}',
            timeout=20,
        )
        response.raise_for_status()
        images.append(base64.b64encode(response.content).decode())
    return images


def _vision_content(images, prompt):
    content = []
    for index, image_b64 in enumerate(images, 1):
        content.extend([
            {"type": "text", "text": f"Image {index} of {len(images)}:"},
            {"type": "image", "source": {
                "type": "base64", "media_type": "image/jpeg", "data": image_b64,
            }},
        ])
    content.append({"type": "text", "text": prompt})
    return content


def _merge_verified_contacts(info, verified):
    verified_name = str(verified.get('client_name') or '').strip()
    if verified_name:
        info['client_name'] = verified_name
    if normalize_civility(info.get('client_civility')) == 'M./Mme':
        info['client_civility'] = verified.get(
            'client_civility', info.get('client_civility', 'M./Mme')
        )
    for key in ('phone', 'address'):
        value = str(verified.get(key) or '').strip()
        if value:
            info[key] = value
    verified_email = str(verified.get('email') or '').strip()
    if verified_email and not verified_email.lower().endswith('@metrastructure.ca'):
        info['email'] = verified_email
    return info


def do_extract_many(chat_id, uid, file_ids):
    uid = str(uid)
    try:
        images = _download_telegram_images(file_ids)
        if not images:
            raise RuntimeError('aucune image lisible')
        model_content = _vision_content(
            images, _photo_extraction_prompt(len(images))
        )
        response = client.messages.create(
            model="claude-sonnet-4-6", max_tokens=1600, temperature=0,
            messages=[{"role": "user", "content": model_content}],
        )
        info = _model_json(response)

        # A second, narrow pass prevents dense email screenshots from losing the
        # body-stated full name or phone number to a shorter From display name.
        contact_response = client.messages.create(
            model="claude-sonnet-4-6", max_tokens=450, temperature=0,
            messages=[{"role": "user", "content": _vision_content(
                images, _contact_verification_prompt()
            )}],
        )
        info = _merge_verified_contacts(info, _model_json(contact_response))

        observations = [
            str(item).strip()
            for item in (info.get('image_observations') or [])
            if str(item).strip()
        ]
        description = str(info.get('project_description') or '').strip()
        if observations and not any(obs.casefold() in description.casefold() for obs in observations):
            description = (description + ' Observations visuelles : ' + '; '.join(observations)).strip()

        yr = datetime.now().strftime('%y')
        service = str(info.get('suggested_service') or '').strip()
        price = info.get('suggested_price') or PRICES.get(service, 3200)
        user_data[uid] = {
            'name': normalize_client_name(info.get('client_name', '')),
            'civility': normalize_civility(info.get('client_civility')),
            'phone': str(info.get('phone') or '').strip(),
            'email': str(info.get('email') or '').strip(),
            'addr': str(info.get('address') or '').strip(),
            'desc': description,
            'image_observations': observations,
            'source_image_count': len(images),
            'property_type': str(info.get('property_type') or '').strip(),
            'service': service,
            'price': price,
            'odsNum': f"ODS{yr}-{random.randint(100,999)}",
            'date': datetime.now().strftime('%Y-%m-%d'),
        }
        save_user_data()
        ask_next_missing(chat_id, uid)
    except Exception as e:
        import traceback
        logger.error(f"Extract error: {e}")
        logger.error(traceback.format_exc())
        current = user_data.get(uid, {})
        current.pop('extracting', None)
        user_data[uid] = current
        save_user_data()
        tg(chat_id, f"❌ Erreur: {str(e)}")


def do_extract(chat_id, uid, file_id):
    """Backward-compatible single-image extraction entry point."""
    return do_extract_many(chat_id, uid, [file_id])


def _finish_photo_batch(chat_id, uid, generation):
    time.sleep(PHOTO_BATCH_DELAY_SECONDS)
    with photo_batch_lock:
        batch = photo_batches.get(uid)
        if not batch or batch['generation'] != generation:
            return
        file_ids = list(batch['file_ids'])[:MAX_EXTRACTION_IMAGES]
        photo_batches.pop(uid, None)
    current = user_data.get(uid, {})
    current['extracting'] = True
    user_data[uid] = current
    save_user_data()
    tg(chat_id, f"🔍 Analyse de {len(file_ids)} image(s) ensemble...")
    do_extract_many(chat_id, uid, file_ids)


def queue_photo_extraction(chat_id, uid, file_id):
    """Collect consecutive Telegram photos before running one combined analysis."""
    uid = str(uid)
    with photo_batch_lock:
        batch = photo_batches.setdefault(uid, {'file_ids': [], 'generation': 0})
        if file_id not in batch['file_ids'] and len(batch['file_ids']) < MAX_EXTRACTION_IMAGES:
            batch['file_ids'].append(file_id)
        batch['generation'] += 1
        generation = batch['generation']
        count = len(batch['file_ids'])
    if count == 1:
        tg(chat_id, "📷 Image reçue. Vous pouvez envoyer les autres images maintenant.")
    executor.submit(_finish_photo_batch, chat_id, uid, generation)

def do_extract_text(chat_id, uid, client_text):
    uid = str(uid)
    try:
        PROMPT = (
            "Extract only information explicitly present in this client email/text. "
            "Do not infer or invent missing information. Return ONLY JSON with these keys: "
            "client_name, client_civility, phone, email, address, soumission_ref, project_description, property_type, suggested_service, suggested_price. "
            "Read the full header and body. If the body explicitly identifies the sender by name, that full name overrides a shortened or different From display name. "
            "Use the From email, never To, Cc, or an @metrastructure.ca address. Extract any phone number written in the body or signature. "
            "project_description must preserve the specific existing condition, requested opinion/report/design, and any explicitly requested tests or deliverables. "
            "For client_civility, return M. for a confidently male first name, Mme for a confidently female first name, "
            "and M./Mme when the name is ambiguous or confidence is low. Never guess weakly. "
            "For address, extract only the components provided by the client. "
            "Do not infer a postal code, city, province, or country. "
            "suggested_service must be one of: Analyse structurale generale, Inspection et rapport structural, "
            "Avis d expert stabilisation et renforcement, Enlevement de mur porteur, Inspection des fondations, "
            "Evaluation des fissures et desordres structuraux, Mur de soutenement, Conception structurale complete, "
            "Analyse structurale sous-sol et ajout au-dessus du garage, Reamenagement interieur avec modification structurale. "
            "suggested_price: integer CAD. ONLY JSON, no markdown.\n\nText:\n"
        ) + client_text

        response = client.messages.create(
            model="claude-sonnet-4-6", max_tokens=800,
            messages=[{"role": "user", "content": PROMPT}]
        )
        result = ''.join(b.text for b in response.content if hasattr(b, 'text'))
        result = result.replace('```json', '').replace('```', '').strip()
        info = json.loads(result)

        yr = datetime.now().strftime('%y')
        ods_num = "ODS" + yr + "-" + str(random.randint(100, 999))
        price = info.get('suggested_price') or PRICES.get(info.get('suggested_service', ''), 3200)

        user_data[uid] = {
            'name': normalize_client_name(info.get('client_name', '')),
            'civility': normalize_civility(info.get('client_civility')),
            'phone': info.get('phone', ''),
            'email': info.get('email', ''),
            'addr': info.get('address', ''),
            'desc': info.get('project_description', ''),
            'service': info.get('suggested_service', ''),
            'price': price,
            'odsNum': ods_num,
            'date': datetime.now().strftime('%Y-%m-%d'),
        }
        save_user_data()
        ask_next_missing(chat_id, uid)
    except Exception as e:
        import traceback
        logger.error("do_extract_text error: " + str(e))
        logger.error(traceback.format_exc())
        tg(chat_id, "❌ Erreur extraction: " + str(e))

def do_excel(chat_id, uid):
    uid = str(uid)
    d = user_data.get(uid)
    logger.info(f"do_excel called: uid={uid}, data={d is not None}")
    if not d:
        tg(chat_id, "❌ Session expirée. Envoyez une nouvelle photo.")
        return
    try:
        logger.info(f"Generating Excel for {d.get('name','?')}")
        tg(chat_id, "⏳ Génération Excel...")
        buf = generate_excel(d)
        fname = "{}_{}. xlsx".format(d.get('odsNum','ODS'), (d.get('name') or 'client').replace(' ','-')).replace(' ','')
        logger.info(f"Excel generated, sending {fname}")
        tg_doc(chat_id, buf, fname, f"✅ Excel — ${d['price']:,} CAD")
    except Exception as e:
        import traceback
        logger.error(f"do_excel error: {e}")
        logger.error(traceback.format_exc())
        tg(chat_id, f"❌ Erreur Excel: {str(e)}")

def valid_client_email(value):
    email = str(value or '').strip()
    return email if re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email) else ''


def build_email_preview(data):
    recipient = valid_client_email(data.get('email'))
    name = normalize_client_name(data.get('name'))
    last_name = name.split()[-1] if name else ''
    civility = normalize_civility(data.get('civility'))
    greeting = (
        f"Bonjour {civility} {last_name},"
        if civility in ('M.', 'Mme') and last_name
        else f"Bonjour {name}," if name else "Bonjour,"
    )
    ods_num = str(data.get('odsNum') or 'ODS').strip()
    # odsNum may also contain the long generated filename. Keep the email
    # subject limited to the official ODS reference.
    ods_match = re.search(r'ODS\d{2}-\d{3}-[A-Z]{3}', ods_num, re.IGNORECASE)
    ods_reference = ods_match.group(0).upper() if ods_match else ods_num.split('-')[0]
    project = str(data.get('project_title') or data.get('service') or 'votre projet').strip()
    address = client_contact_fields(data)['address']
    subject = f"{ods_reference} – Offre de service"
    body = (
        f"{greeting}\n\n"
        f"Veuillez trouver ci-joint notre offre de service {ods_num} concernant "
        f"{project.lower()} pour le projet situé au {address}.\n\n"
        "N'hésitez pas à nous contacter pour toute question.\n\n"
        "Cordialement,\n\n"
        "Arash Rohani, ing., P.Eng.\n"
        "Président – Ingénieur en structure\n"
        f"{DISPLAY_BRAND}\n"
        "arash.rohani@metrastructure.ca | (438) 867-4131"
    )
    return recipient, subject, body


def email_body_html(body):
    """Render the approved email text with a consistent Metra signature."""
    paragraphs = [
        html.escape(part).replace('\n', '<br>')
        for part in body.rsplit('Cordialement,', 1)[0].strip().split('\n\n')
    ]
    content = ''.join(f'<p>{paragraph}</p>' for paragraph in paragraphs)
    return (
        '<div style="font-family:Arial,Helvetica,sans-serif;font-size:11pt;'
        'line-height:1.45;color:#1f1f1f">'
        f'{content}'
        '<p>Cordialement,</p>'
        '<div style="border-left:4px solid #f5a623;padding-left:12px;margin-top:16px">'
        '<strong style="font-size:12pt;color:#102a43">Arash Rohani, ing., P.Eng.</strong><br>'
        'Président – Ingénieur en structure<br>'
        f'<strong>{DISPLAY_BRAND}</strong><br>'
        '<a href="mailto:arash.rohani@metrastructure.ca" style="color:#1155cc">'
        'arash.rohani@metrastructure.ca</a> | '
        '<a href="tel:+14388674131" style="color:#1155cc">(438) 867-4131</a><br>'
        '<a href="https://metrastructure.ca" style="color:#1155cc">metrastructure.ca</a>'
        '</div></div>'
    )


def microsoft_email_config():
    """Read Railway variables at send time and report missing names without secrets."""
    config = {
        'MS_TENANT_ID': str(
            os.environ.get('MS_TENANT_ID') or MS_TENANT_ID or ''
        ).strip(),
        'MS_CLIENT_ID': str(
            os.environ.get('MS_CLIENT_ID') or MS_CLIENT_ID or ''
        ).strip(),
        'MS_CLIENT_SECRET': str(os.environ.get('MS_CLIENT_SECRET') or '').strip(),
        'EMAIL_SENDER': str(
            os.environ.get('EMAIL_SENDER') or EMAIL_SENDER or ''
        ).strip(),
    }
    missing = [name for name, value in config.items() if not value]
    if missing:
        logger.error(
            "Microsoft 365 missing environment variables: %s",
            ", ".join(missing),
        )
        raise RuntimeError(
            "Configuration Microsoft 365 incomplète : " + ", ".join(missing)
        )
    return config


def graph_access_token():
    config = microsoft_email_config()
    response = req.post(
        f"https://login.microsoftonline.com/{config['MS_TENANT_ID']}/oauth2/v2.0/token",
        data={
            'client_id': config['MS_CLIENT_ID'],
            'client_secret': config['MS_CLIENT_SECRET'],
            'scope': 'https://graph.microsoft.com/.default',
            'grant_type': 'client_credentials',
        },
        timeout=20,
    )
    if response.status_code != 200:
        logger.error("Microsoft token error: %s %s", response.status_code, response.text[:300])
        raise RuntimeError("Authentification Microsoft 365 impossible.")
    return response.json()['access_token']


def safe_archive_filename(value):
    """Return a OneDrive-safe filename while preserving French characters."""
    cleaned = re.sub(r'[<>:"/\\|?*#%]+', '-', str(value or 'ODS'))
    cleaned = re.sub(r'\s+', ' ', cleaned).strip(' .-')
    return cleaned[:180] or 'ODS'


def upload_onedrive_path(token, sender, relative_path, content, content_type):
    """Upload or replace one file at a path in the sender's OneDrive."""
    encoded_path = urllib.parse.quote(relative_path, safe='/')
    encoded_sender = urllib.parse.quote(sender, safe='')
    response = req.put(
        f"https://graph.microsoft.com/v1.0/users/{encoded_sender}/drive/"
        f"root:/{encoded_path}:/content",
        headers={
            'Authorization': f'Bearer {token}',
            'Content-Type': content_type,
        },
        data=content,
        timeout=60,
    )
    if response.status_code not in (200, 201):
        logger.error(
            "OneDrive upload error for %s: %s %s",
            relative_path,
            response.status_code,
            response.text[:500],
        )
        raise RuntimeError(f"échec de l'archivage de {relative_path}")


def download_onedrive_path(token, sender, relative_path):
    """Download a file from the sender's OneDrive."""
    encoded_path = urllib.parse.quote(relative_path, safe='/')
    encoded_sender = urllib.parse.quote(sender, safe='')
    response = req.get(
        f"https://graph.microsoft.com/v1.0/users/{encoded_sender}/drive/"
        f"root:/{encoded_path}:/content",
        headers={'Authorization': f'Bearer {token}'},
        timeout=60,
    )
    if response.status_code != 200:
        logger.error(
            "OneDrive download error for %s: %s %s",
            relative_path,
            response.status_code,
            response.text[:500],
        )
        raise RuntimeError(f"impossible de télécharger {relative_path}")
    return response.content


ODS_LIST_HEADERS = {
    'no': 'No',
    'year': 'Year ',
    'month': 'Month',
    'description': 'Description',
    'price': 'Price($)',
    'date': 'Date',
    'status': 'Status',
    'accepted_price': 'Accepted Price',
    'source': 'Source',
    'contact': 'Contact',
    'accepted_at': 'Date of acceptation',
    'email': 'Email',
}


def _ods_event_datetime(data, accepted_at=None):
    value = accepted_at or data.get('email_sent_at') or data.get('date')
    if isinstance(value, datetime):
        return value
    if value:
        try:
            return datetime.fromisoformat(str(value).replace('Z', '+00:00')).replace(tzinfo=None)
        except ValueError:
            pass
    return datetime.now()


def _copy_ods_list_row_style(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy.copy(source.alignment)
        if source.protection:
            target.protection = copy.copy(source.protection)
    if source_row in ws.row_dimensions:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def _enable_workbook_recalculation(workbook):
    """Keep List.xlsx saveable even when its calc properties are missing."""
    if getattr(workbook, 'calculation', None) is None:
        workbook.calculation = CalcProperties(
            calcMode='auto',
            fullCalcOnLoad=True,
            forceFullCalc=True,
        )
        return
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = 'auto'


def upsert_ods_list_workbook(workbook, data, status='In process', accepted_at=None):
    """Insert or update one ODS row, keyed by the full ODS reference."""
    event_date = _ods_event_datetime(data, accepted_at)
    reference = offer_reference(data)
    reference_year = re.search(r'^ODS(\d{2})-', reference, re.I)
    sheet_year = int(f"20{reference_year.group(1)}") if reference_year else event_date.year
    sheet_name = f"data {sheet_year}"
    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(f"onglet {sheet_name} introuvable dans List.xlsx")
    ws = workbook[sheet_name]
    headers = {
        str(cell.value or '').strip(): cell.column
        for cell in ws[1]
        if cell.value is not None
    }
    columns = {}
    for key, label in ODS_LIST_HEADERS.items():
        col = headers.get(label.strip())
        if not col:
            raise RuntimeError(f"colonne {label} introuvable dans List.xlsx")
        columns[key] = col

    target_row = None
    last_data_row = 1
    highest_no = 0
    for row in range(2, ws.max_row + 1):
        number = ws.cell(row, columns['no']).value
        description = str(ws.cell(row, columns['description']).value or '')
        if number not in (None, '') or description:
            last_data_row = row
        try:
            highest_no = max(highest_no, int(number or 0))
        except (TypeError, ValueError):
            pass
        if reference and reference in description.upper():
            target_row = row
            break

    is_new = target_row is None
    if is_new:
        target_row = last_data_row + 1
        _copy_ods_list_row_style(ws, last_data_row, target_row)
        ws.cell(target_row, columns['no']).value = highest_no + 1

    price = float(data.get('price') or 0)
    accepted = status == 'Accept'
    sent_date = _ods_event_datetime(data)
    ws.cell(target_row, columns['year']).value = sheet_year
    ws.cell(target_row, columns['month']).value = sent_date.strftime('%B')
    ws.cell(target_row, columns['description']).value = str(data.get('odsNum') or reference)
    ws.cell(target_row, columns['price']).value = price
    ws.cell(target_row, columns['date']).value = sent_date
    ws.cell(target_row, columns['status']).value = status
    ws.cell(target_row, columns['accepted_price']).value = price if accepted else 0
    source = data.get('source') or data.get('lead_source') or data.get('referral_source')
    contact = normalize_client_name(data.get('name'))
    email = str(data.get('email') or '').strip()
    if is_new or source:
        ws.cell(target_row, columns['source']).value = source or ''
    if is_new or contact:
        ws.cell(target_row, columns['contact']).value = contact
    ws.cell(target_row, columns['accepted_at']).value = event_date if accepted else None
    if is_new or email:
        ws.cell(target_row, columns['email']).value = email
    ws.cell(target_row, columns['date']).number_format = 'yyyy-mm-dd'
    ws.cell(target_row, columns['accepted_at']).number_format = 'yyyy-mm-dd'

    for table in ws.tables.values():
        start, _ = table.ref.split(':', 1)
        end_col = openpyxl.utils.get_column_letter(ws.max_column)
        table.ref = f"{start}:{end_col}{max(target_row, last_data_row)}"
    _enable_workbook_recalculation(workbook)
    return target_row


def sync_ods_list(data, status='In process', accepted_at=None):
    """Safely synchronize an ODS status with OneDrive List.xlsx."""
    with ods_list_lock:
        config = microsoft_email_config()
        token = graph_access_token()
        sender = config['EMAIL_SENDER']
        content = download_onedrive_path(token, sender, ODS_LIST_PATH)
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        row = upsert_ods_list_workbook(workbook, data, status, accepted_at)
        output = io.BytesIO()
        workbook.save(output)
        upload_onedrive_path(
            token,
            sender,
            ODS_LIST_PATH,
            output.getvalue(),
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        logger.info("ODS List synchronized: %s status=%s row=%s", offer_reference(data), status, row)
        return row


def upload_onedrive_file(token, sender, filename, content, content_type):
    """Upload or replace one file in the approved ODS archive folder."""
    upload_onedrive_path(
        token,
        sender,
        f"Metra Structure Inc/Offre de service/{filename}",
        content,
        content_type,
    )


def archive_ods_files(data, token, sender, pdf_bytes):
    """Archive the final PDF and Excel in Metra's OneDrive ODS folder."""
    base_name = safe_archive_filename(data.get('odsNum') or 'ODS')
    excel = generate_excel(data)
    excel.seek(0)
    upload_onedrive_file(
        token,
        sender,
        f"{base_name}.pdf",
        pdf_bytes,
        'application/pdf',
    )
    upload_onedrive_file(
        token,
        sender,
        f"{base_name}.xlsx",
        excel.read(),
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    return [f"{base_name}.pdf", f"{base_name}.xlsx"]


PROJECT_SUBFOLDERS = [
    'Approvals & Permits',
    'Calculations',
    'Correspondence',
    'Drawings Plans',
    'Reports',
    'Specifications',
]


def onedrive_user_url(sender, suffix):
    encoded_sender = urllib.parse.quote(sender, safe='')
    return f"https://graph.microsoft.com/v1.0/users/{encoded_sender}/drive/{suffix}"


def list_onedrive_children(token, sender, parent_path):
    """List every direct child of a OneDrive folder."""
    encoded_path = urllib.parse.quote(parent_path, safe='/')
    url = onedrive_user_url(sender, f"root:/{encoded_path}:/children")
    headers = {'Authorization': f'Bearer {token}'}
    items = []
    params = {'$select': 'name,folder,webUrl', '$top': '999'}
    while url:
        response = req.get(url, headers=headers, params=params, timeout=30)
        if response.status_code != 200:
            logger.error(
                "OneDrive list error for %s: %s %s",
                parent_path,
                response.status_code,
                response.text[:500],
            )
            raise RuntimeError("impossible de lire les numéros de projet OneDrive")
        payload = response.json()
        items.extend(payload.get('value', []))
        url = payload.get('@odata.nextLink')
        params = None
    return items


def get_onedrive_item(token, sender, relative_path):
    encoded_path = urllib.parse.quote(relative_path, safe='/')
    response = req.get(
        onedrive_user_url(sender, f"root:/{encoded_path}"),
        headers={'Authorization': f'Bearer {token}'},
        params={'$select': 'name,folder,webUrl'},
        timeout=30,
    )
    if response.status_code != 200:
        raise RuntimeError(f"dossier OneDrive introuvable : {relative_path}")
    return response.json()


def create_onedrive_folder(token, sender, parent_path, folder_name):
    """Create a folder idempotently and return its metadata."""
    encoded_parent = urllib.parse.quote(parent_path, safe='/')
    response = req.post(
        onedrive_user_url(sender, f"root:/{encoded_parent}:/children"),
        headers={
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json',
        },
        json={
            'name': folder_name,
            'folder': {},
            '@microsoft.graph.conflictBehavior': 'fail',
        },
        timeout=30,
    )
    if response.status_code in (200, 201):
        return response.json()
    if response.status_code == 409:
        return get_onedrive_item(token, sender, f"{parent_path}/{folder_name}")
    logger.error(
        "OneDrive folder error for %s/%s: %s %s",
        parent_path,
        folder_name,
        response.status_code,
        response.text[:500],
    )
    raise RuntimeError(f"impossible de créer le dossier {folder_name}")


def project_year_and_code(data):
    ods = str(data.get('odsNum') or '')
    match = re.search(r'ODS(\d{2})-\d{3}-([A-Z]{3})', ods, re.IGNORECASE)
    if match:
        return f"20{match.group(1)}", match.group(2).upper()
    return datetime.now().strftime('%Y'), str(data.get('file_code') or 'PRJ').upper()[:3]


def next_project_number(items, year):
    prefix = year[-2:]
    numbers = []
    for item in items:
        if 'folder' not in item:
            continue
        match = re.match(rf'^P{prefix}-(\d{{3}})(?:-|$)', str(item.get('name') or ''), re.I)
        if match:
            numbers.append(int(match.group(1)))
    return max(numbers, default=0) + 1


def project_folder_name(data, number, year, code):
    title = str(data.get('project_title') or data.get('service') or 'Projet').strip()
    title = re.sub(r'[\s–—]+', '-', safe_archive_filename(title))
    title = re.sub(r'-+', '-', title).strip('-')[:110] or 'Projet'
    return f"P{year[-2:]}-{number:03d}-{code}-{title}"


def create_project_from_ods(data):
    """Create the numbered project structure and place ODS files in Correspondence."""
    config = microsoft_email_config()
    token = graph_access_token()
    sender = config['EMAIL_SENDER']
    year, code = project_year_and_code(data)
    projects_root = 'Metra Structure Inc/Projects'
    create_onedrive_folder(token, sender, projects_root, year)
    year_root = f"{projects_root}/{year}"

    folder_name = str(data.get('project_folder') or '').strip()
    if folder_name:
        project_item = create_onedrive_folder(token, sender, year_root, folder_name)
    else:
        items = list_onedrive_children(token, sender, year_root)
        number = next_project_number(items, year)
        folder_name = project_folder_name(data, number, year, code)
        project_item = create_onedrive_folder(token, sender, year_root, folder_name)
        data['project_folder'] = folder_name

    project_path = f"{year_root}/{folder_name}"
    for subfolder in PROJECT_SUBFOLDERS:
        create_onedrive_folder(token, sender, project_path, subfolder)

    base_name = safe_archive_filename(data.get('odsNum') or 'ODS')
    pdf = generate_pdf(data)
    pdf.seek(0)
    excel = generate_excel(data)
    excel.seek(0)
    correspondence = f"{project_path}/Correspondence"
    upload_onedrive_path(
        token, sender, f"{correspondence}/{base_name}.pdf",
        pdf.read(), 'application/pdf',
    )
    upload_onedrive_path(
        token, sender, f"{correspondence}/{base_name}.xlsx",
        excel.read(),
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    return folder_name, project_item.get('webUrl', '')


def send_ods_email(data):
    recipient, subject, body = build_email_preview(data)
    if not recipient:
        raise ValueError("Le courriel du client est manquant ou invalide.")
    pdf = generate_pdf(data)
    pdf.seek(0)
    pdf_bytes = pdf.read()
    filename = "{}_{}.pdf".format(
        data.get('odsNum', 'ODS'),
        (data.get('name') or 'client').replace(' ', '-'),
    )
    payload = {
        'message': {
            'subject': subject,
            'body': {'contentType': 'HTML', 'content': email_body_html(body)},
            'toRecipients': [
                {'emailAddress': {'address': recipient}}
            ],
            'attachments': [{
                '@odata.type': '#microsoft.graph.fileAttachment',
                'name': filename,
                'contentType': 'application/pdf',
                'contentBytes': base64.b64encode(pdf_bytes).decode('ascii'),
            }],
        },
        'saveToSentItems': True,
    }
    config = microsoft_email_config()
    token = graph_access_token()
    sender = urllib.parse.quote(config['EMAIL_SENDER'], safe='')
    response = req.post(
        f"https://graph.microsoft.com/v1.0/users/{sender}/sendMail",
        headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'},
        json=payload,
        timeout=45,
    )
    if response.status_code != 202:
        logger.error("Microsoft sendMail error: %s %s", response.status_code, response.text[:500])
        raise RuntimeError("Microsoft 365 a refusé l'envoi du courriel.")
    archive_files = []
    archive_error = None
    try:
        archive_files = archive_ods_files(data, token, config['EMAIL_SENDER'], pdf_bytes)
    except Exception as exc:
        archive_error = str(exc)
        logger.error("OneDrive ODS archive error: %s", exc)
    return recipient, subject, archive_files, archive_error


def show_email_confirmation(chat_id, uid):
    uid = str(uid)
    data = user_data.get(uid, {})
    recipient, subject, body = build_email_preview(data)
    if not recipient:
        tg(chat_id, "⚠️ Aucun courriel client valide. Le PDF n'a pas été envoyé par courriel.")
        return
    data['email_subject'] = subject
    data['email_body'] = body
    data['email_recipient'] = recipient
    data['email_sent_at'] = None
    user_data[uid] = data
    save_user_data()
    preview = (
        "✉️ Aperçu du courriel\n\n"
        f"De : {EMAIL_SENDER}\n"
        f"À : {recipient}\n"
        f"Objet : {subject}\n\n"
        f"{body}\n\n"
        "Pièce jointe : PDF de l'offre de service"
    )
    tg(
        chat_id,
        preview,
        [
            [{'text': '✅ Confirmer et envoyer', 'callback_data': 'email_send'}],
            [{'text': '❌ Ne pas envoyer', 'callback_data': 'email_cancel'}],
        ],
    )


def do_send_email(chat_id, uid):
    uid = str(uid)
    data = user_data.get(uid)
    if not data:
        tg(chat_id, "❌ Session expirée. Générez une nouvelle offre.")
        return
    if data.get('email_sent_at'):
        tg(chat_id, "⚠️ Ce courriel a déjà été envoyé. Aucun second envoi effectué.")
        return
    if data.get('email_sending'):
        tg(chat_id, "⏳ Envoi déjà en cours.")
        return
    data['email_sending'] = True
    user_data[uid] = data
    save_user_data()
    try:
        tg(chat_id, "⏳ Envoi du courriel via Microsoft 365...")
        recipient, subject, archive_files, archive_error = send_ods_email(data)
        data['email_sent_at'] = datetime.now().isoformat(timespec='seconds')
        data['email_sending'] = False
        user_data[uid] = data
        save_user_data()
        record_sent_offer(uid, data)
        list_sync_error = None
        try:
            sync_ods_list(data, 'In process')
        except Exception as exc:
            list_sync_error = str(exc)
            logger.exception("ODS List initial synchronization failed")
        tg(
            chat_id,
            f"✅ Courriel envoyé à {recipient}\n"
            f"Objet : {subject}\n"
            "Une copie est enregistrée dans les éléments envoyés.",
        )
        if archive_error:
            tg(
                chat_id,
                "⚠️ Courriel envoyé, mais archivage OneDrive non complété : "
                + archive_error,
            )
        else:
            tg(
                chat_id,
                "☁️ Archives OneDrive enregistrées :\n"
                + "\n".join(f"• {name}" for name in archive_files),
            )
        if list_sync_error:
            tg(chat_id, "⚠️ Offre envoyée, mais List.xlsx n'a pas été mis à jour : " + list_sync_error)
        else:
            tg(chat_id, "📊 List.xlsx mis à jour : In process")
        tg(
            chat_id,
            "Mettre à jour le statut de cette offre :",
            [
                [{'text': '✅ Convertir en projet', 'callback_data': 'project_create'}],
                [
                    {'text': '❌ Refused', 'callback_data': 'ods_status:Refused'},
                    {'text': '🔒 Closed', 'callback_data': 'ods_status:Closed'},
                ],
                [{'text': '⏸ Hold', 'callback_data': 'ods_status:Hold'}],
            ],
        )
    except Exception as exc:
        data['email_sending'] = False
        user_data[uid] = data
        save_user_data()
        logger.error("ODS email error: %s", exc)
        tg(chat_id, f"❌ Envoi impossible : {exc}")



def next_invoice_number(items, year=None):
    year = year or datetime.now().strftime('%y')
    numbers = []
    for item in items:
        name = str(item.get('name') or '')
        match = re.search(rf'FAC\s+P{year}-(\d{{1,4}})(?:-|\b)', name, re.I)
        if not match:
            match = re.search(r'(?:facture|invoice|FAC)[^0-9]*(\d{1,4})', name, re.I)
        if match:
            numbers.append(int(match.group(1)))
    return max(numbers, default=0) + 1


def invoice_email_html(data, invoice_number, total, due_date):
    client = html.escape(str(data.get('name') or 'Client'))
    project = html.escape(str(data.get('project_folder') or data.get('odsNum') or ''))
    return (
        f"<p>Bonjour {client},</p>"
        f"<p>Veuillez trouver ci-joint la facture no {invoice_number} "
        f"relative au projet <b>{project}</b>.</p>"
        f"<p>Montant total : <b>{total:,.2f} $ CAD</b><br>"
        f"Échéance : <b>{due_date.isoformat()}</b></p>"
        "<p>N'hésitez pas à nous contacter pour toute question.</p>"
        "<p>Cordialement,<br><b>Service de comptabilité</b><br>"
        f"{DISPLAY_BRAND}<br>"
        "accounting@metrastructure.ca</p>"
    )


def send_invoice_email(token, recipient, filename, pdf_bytes, data, invoice_number, values, due_date):
    if not valid_client_email(recipient):
        raise ValueError("Le courriel du client est manquant ou invalide.")
    sender = urllib.parse.quote(INVOICE_EMAIL_SENDER, safe='')
    subject = f"Facture no {invoice_number} - {data.get('project_folder') or data.get('odsNum') or 'Projet'}"
    payload = {
        'message': {
            'subject': subject,
            'body': {
                'contentType': 'HTML',
                'content': invoice_email_html(
                    data, invoice_number, values['total'], due_date
                ),
            },
            'toRecipients': [{'emailAddress': {'address': recipient}}],
            'ccRecipients': [
                {'emailAddress': {'address': INVOICE_CC_EMAIL}}
            ],
            'attachments': [{
                '@odata.type': '#microsoft.graph.fileAttachment',
                'name': filename,
                'contentType': 'application/pdf',
                'contentBytes': base64.b64encode(pdf_bytes).decode('ascii'),
            }],
        },
        'saveToSentItems': True,
    }
    response = req.post(
        f"https://graph.microsoft.com/v1.0/users/{sender}/sendMail",
        headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'},
        json=payload,
        timeout=45,
    )
    if response.status_code != 202:
        logger.error("Invoice sendMail error: %s %s", response.status_code, response.text[:500])
        raise RuntimeError("Microsoft 365 a refusé l'envoi depuis accounting.")


def show_invoice_options(chat_id, uid):
    data = user_data.get(str(uid), {})
    if not data.get('project_created') or not data.get('project_folder'):
        tg(chat_id, "⚠️ Convertissez d'abord l'offre en projet.")
        return
    if data.get('invoice_issued_at'):
        tg(
            chat_id,
            f"⚠️ La facture no {data.get('invoice_number')} a déjà été émise pour cette étape.",
        )
        return
    tg(
        chat_id,
        "🧾 Facture initiale\n\nChoisissez le mode de facturation :",
        [
            [{'text': '✅ 25 % du contrat', 'callback_data': 'invoice_pct:25'}],
            [{'text': '📊 Autre pourcentage', 'callback_data': 'invoice_pct_other'}],
            [{'text': '💵 Montant fixe', 'callback_data': 'invoice_fixed'}],
            [{'text': '❌ Plus tard', 'callback_data': 'invoice_cancel'}],
        ],
    )


def show_invoice_preview(chat_id, uid, percentage=None, fixed_amount=None):
    uid = str(uid)
    data = user_data.get(uid, {})
    values = invoice_values(data.get('price') or 0, percentage, fixed_amount)
    data['pending_invoice'] = {
        'percentage': percentage,
        'fixed_amount': fixed_amount,
    }
    user_data[uid] = data
    save_user_data()
    quantity = (
        f"{float(percentage):g} % du contrat"
        if percentage is not None else "Montant fixe"
    )
    tg(
        chat_id,
        "🧾 Aperçu de la facture\n\n"
        f"Projet : {data.get('project_folder')}\n"
        f"Client : {data.get('name')}\n"
        f"Méthode : {quantity}\n"
        f"Contrat : {values['contract']:,.2f} $\n"
        f"Sous-total : {values['subtotal']:,.2f} $\n"
        f"TPS : {values['gst']:,.2f} $\n"
        f"TVQ : {values['qst']:,.2f} $\n"
        f"Total : {values['total']:,.2f} $\n\n"
        "Le prochain numéro disponible sera vérifié au moment de l'émission.",
        [
            [{'text': '✅ Confirmer et envoyer', 'callback_data': 'invoice_confirm'}],
            [{'text': '✏️ Modifier', 'callback_data': 'invoice_start'}],
            [{'text': '❌ Annuler', 'callback_data': 'invoice_cancel'}],
        ],
    )


def do_issue_invoice(chat_id, uid):
    uid = str(uid)
    if not invoice_creation_lock.acquire(blocking=False):
        tg(chat_id, "⏳ Une facture est déjà en cours de création.")
        return
    try:
        data = user_data.get(uid, {})
        pending = data.get('pending_invoice') or {}
        missing_client_details = [
            label for field, label in (
                ('name', 'nom du client'),
                ('addr', 'adresse'),
                ('phone', 'téléphone'),
                ('email', 'courriel'),
            )
            if not str(data.get(field) or '').strip()
        ]
        if missing_client_details:
            tg(
                chat_id,
                "⛔ Facture non envoyée : informations client incomplètes.\n"
                "À compléter : " + ", ".join(missing_client_details) + ".\n\n"
                "Rouvrez Facturation et sélectionnez le projet pour les compléter.",
            )
            return
        if not valid_client_email(data.get('email')):
            tg(chat_id, "⛔ Facture non envoyée : adresse courriel client invalide.")
            return
        if not data.get('project_folder'):
            tg(chat_id, "❌ Numéro de projet introuvable.")
            return
        if data.get('invoice_issued_at'):
            tg(chat_id, f"⚠️ Facture déjà émise : no {data.get('invoice_number')}.")
            return
        tg(chat_id, "⏳ Vérification du numéro et création de la facture...")
        token = graph_access_token()
        items = list_onedrive_children(token, INVOICE_DRIVE_OWNER, INVOICE_FOLDER)
        number = next_invoice_number(items)
        pdf_bytes, values, due_date = generate_invoice_pdf(
            data,
            number,
            percentage=pending.get('percentage'),
            fixed_amount=pending.get('fixed_amount'),
            logo_path=LOGOS['metra'],
        )
        filename = invoice_filename(number, data)
        upload_onedrive_path(
            token, INVOICE_DRIVE_OWNER, f"{INVOICE_FOLDER}/{filename}",
            pdf_bytes, 'application/pdf',
        )
        project_path = (
            f"Metra Structure Inc/Projects/{datetime.now().strftime('%Y')}/"
            f"{data['project_folder']}/Correspondence/{filename}"
        )
        project_archive_error = None
        try:
            upload_onedrive_path(
                token, INVOICE_DRIVE_OWNER, project_path,
                pdf_bytes, 'application/pdf',
            )
        except Exception as exc:
            project_archive_error = str(exc)
            logger.warning("Project invoice copy failed: %s", exc)

        recipient = valid_client_email(data.get('email'))
        send_invoice_email(
            token, recipient, filename, pdf_bytes, data, number, values, due_date
        )
        data.update({
            'invoice_number': number,
            'invoice_filename': filename,
            'invoice_subtotal': values['subtotal'],
            'invoice_total': values['total'],
            'invoice_due_date': due_date.isoformat(),
            'invoice_issued_at': datetime.now().isoformat(timespec='seconds'),
            'pending_invoice': None,
        })
        user_data[uid] = data
        save_user_data()
        record_sent_offer(uid, data)
        tg(
            chat_id,
            f"✅ Facture no {number} créée et envoyée à {recipient}\n"
            f"CC : {INVOICE_CC_EMAIL}\n"
            f"Fichier : {filename}\n"
            f"Total : {values['total']:,.2f} $\n"
            f"Échéance : {due_date.isoformat()}",
        )
        if project_archive_error:
            tg(chat_id, "⚠️ Copie classée dans Financial, mais pas dans le dossier projet : " + project_archive_error)
        else:
            tg(chat_id, "☁️ Facture classée dans Financial et dans le dossier du projet.")
    except Exception as exc:
        logger.exception("Invoice generation failed")
        tg(chat_id, f"❌ Création de la facture impossible : {exc}")
    finally:
        invoice_creation_lock.release()


def do_create_project(chat_id, uid, offer_ref=None):
    if not project_creation_lock.acquire(blocking=False):
        tg(chat_id, "⏳ Une création de projet est déjà en cours. Réessayez dans un instant.")
        return
    try:
        return _do_create_project(chat_id, uid, offer_ref)
    finally:
        project_creation_lock.release()


def do_update_ods_status(chat_id, uid, status, offer_ref=None):
    allowed = {'Refused', 'Closed', 'Hold', 'In process'}
    if status not in allowed:
        tg(chat_id, "❌ Statut non valide.")
        return
    uid = str(uid)
    history_record = offers_history.get(uid, {}).get(offer_ref) if offer_ref else None
    data = (history_record or {}).get('data') if history_record else user_data.get(uid)
    if not data or not data.get('odsNum'):
        tg(chat_id, "❌ Offre introuvable dans la session active.")
        return
    try:
        sync_ods_list(data, status)
        if history_record is not None:
            history_record['status'] = status
            save_offers_history()
        tg(chat_id, f"📊 List.xlsx mis à jour : {status}")
    except Exception as exc:
        logger.exception("ODS List status synchronization failed")
        tg(chat_id, f"❌ Mise à jour de List.xlsx impossible : {exc}")


def _do_create_project(chat_id, uid, offer_ref=None):
    uid = str(uid)
    history_record = None
    if offer_ref:
        history_record = offers_history.get(uid, {}).get(offer_ref)
        data = (history_record or {}).get('data')
    else:
        data = user_data.get(uid)
    if not data:
        tg(chat_id, "❌ Offre introuvable. Ouvrez de nouveau la liste.")
        return
    if not data.get('email_sent_at'):
        tg(chat_id, "⚠️ Envoyez d'abord l'offre au client.")
        return
    if data.get('project_created') or (history_record and history_record.get('converted')):
        link = (history_record or {}).get('project_web_url') or data.get('project_web_url') or ''
        folder = (history_record or {}).get('project_folder') or data.get('project_folder')
        message = f"✅ Projet déjà créé : {folder}"
        if link:
            message += f"\n{link}"
        user_data[uid] = data
        save_user_data()
        tg(
            chat_id,
            message,
            [[{'text': '🧾 Créer la facture initiale', 'callback_data': 'invoice_start'}]],
        )
        return
    if not offer_ref and data.get('project_creating'):
        tg(chat_id, "⏳ Création du projet déjà en cours.")
        return

    data['project_creating'] = True
    if history_record is not None:
        history_record['data'] = history_data_copy(data)
        save_offers_history()
    else:
        user_data[uid] = data
        save_user_data()
    try:
        tg(chat_id, "⏳ Création du dossier de projet dans OneDrive...")
        folder_name, web_url = create_project_from_ods(data)
        accepted_at = datetime.now()
        list_sync_error = None
        try:
            sync_ods_list(data, 'Accept', accepted_at=accepted_at)
        except Exception as exc:
            list_sync_error = str(exc)
            logger.exception("ODS List acceptance synchronization failed")
        data['project_created'] = True
        data['project_creating'] = False
        data['project_web_url'] = web_url
        if history_record is not None:
            history_record.update({
                'data': history_data_copy(data),
                'converted': True,
                'project_folder': folder_name,
                'project_web_url': web_url,
                'status': 'Accept',
            })
            save_offers_history()
            current = user_data.get(uid)
            if current and offer_reference(current) == offer_ref:
                current.update({
                    'project_created': True,
                    'project_creating': False,
                    'project_folder': folder_name,
                    'project_web_url': web_url,
                })
                save_user_data()
        else:
            user_data[uid] = data
            save_user_data()
            record_sent_offer(uid, data)
        user_data[uid] = data
        save_user_data()
        message = (
            f"✅ Projet créé : {folder_name}\n"
            "6 sous-dossiers créés.\n"
            "PDF et Excel classés dans Correspondence."
        )
        if web_url:
            message += f"\n\n🔗 {web_url}"
        if list_sync_error:
            message += f"\n\n⚠️ Projet créé, mais List.xlsx non mis à jour : {list_sync_error}"
        else:
            message += "\n📊 List.xlsx mis à jour : Accept"
        tg(
            chat_id,
            message,
            [[{'text': '🧾 Créer la facture initiale', 'callback_data': 'invoice_start'}]],
        )
    except Exception as exc:
        data['project_creating'] = False
        if history_record is not None:
            history_record['data'] = history_data_copy(data)
            save_offers_history()
        else:
            user_data[uid] = data
            save_user_data()
        logger.error("Project creation error: %s", exc)
        tg(chat_id, f"❌ Création du projet impossible : {exc}")


def do_pdf(chat_id, uid):
    uid = str(uid)
    d = user_data.get(uid)
    logger.info(f"do_pdf called: uid={uid}, data={d is not None}")
    if not d:
        tg(chat_id, "❌ Session expirée. Envoyez une nouvelle photo.")
        return
    try:
        logger.info(f"Generating PDF for {d.get('name','?')}")
        tg(chat_id, "⏳ Génération PDF...")
        buf = generate_pdf(d)
        fname = "{}_{}.pdf".format(
            d.get('odsNum', 'ODS'),
            (d.get('name') or 'client').replace(' ', '-'),
        )
        logger.info(f"PDF generated, sending {fname}")
        tg_doc(chat_id, buf, fname, f"✅ PDF — ${d['price']:,} CAD")
        show_email_confirmation(chat_id, uid)
    except Exception as e:
        import traceback
        logger.error(f"do_pdf error: {e}")
        logger.error(traceback.format_exc())
        tg(chat_id, f"❌ Erreur PDF: {str(e)}")

@app.route('/')
def index():
    with open(os.path.join(BASE, 'index.html'), 'r', encoding='utf-8') as f:
        return Response(f.read(), mimetype='text/html')

def has_setup_access():
    supplied = request.headers.get('X-Setup-Secret', '') or request.args.get('key', '')
    return bool(SETUP_SECRET) and hmac.compare_digest(supplied, SETUP_SECRET)


@app.route('/api/extract', methods=['POST'])
def api_extract():
    if not has_setup_access():
        return jsonify({'error': 'forbidden'}), 403
    data = request.json
    prompt = """Extract client info. Return ONLY JSON: {"client_name":"","phone":"","email":"","address":"","soumission_ref":"","project_description":"","property_type":"","suggested_service":"","suggested_price":0}"""
    response = client.messages.create(
        model="claude-sonnet-4-6", max_tokens=800,
        messages=[{"role":"user","content":[
            {"type":"image","source":{"type":"base64","media_type":data.get('mime','image/jpeg'),"data":data.get('image')}},
            {"type":"text","text":prompt}
        ]}]
    )
    text = ''.join(b.text for b in response.content if hasattr(b,'text'))
    return jsonify(json.loads(text.replace('```json','').replace('```','').strip()))

@app.route('/api/generate-pdf', methods=['POST'])
def api_pdf():
    if not has_setup_access():
        return jsonify({'error': 'forbidden'}), 403
    data = request.json
    buf = generate_pdf(data)
    filename = f"{data.get('odsNum','ODS')}_{data.get('name','client').replace(' ','-')}.pdf"
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=filename)

def handle_update(data):
    """Process Telegram update in background thread — webhook returns immediately"""
    try:
        msg = data.get('message', {})
        cb = data.get('callback_query', {})

        actor = cb.get('from') if cb else msg.get('from')
        actor_id = actor.get('id') if actor else None
        chat = cb.get('message', {}).get('chat', {}) if cb else msg.get('chat', {})
        chat_id_for_denial = chat.get('id')
        if actor_id not in ALLOWED_USERS:
            if chat_id_for_denial:
                tg(chat_id_for_denial, "⛔ Ce bot est privé.")
            return

        if cb:
            uid = str(cb['from']['id'])
            chat_id = cb['message']['chat']['id']
            cdata = cb.get('data', '')
            # Answer callback immediately to remove loading spinner
            try:
                req.post(f'https://api.telegram.org/bot{BOT_TOKEN}/answerCallbackQuery',
                         json={'callback_query_id': cb['id']}, timeout=3)
            except:
                pass
            logger.info(f"CB: {cdata} uid={uid} has_data={uid in user_data} keys={list(user_data.keys())}")
            if cdata in ('xl', 'pdf', 'both'):
                if uid not in user_data:
                    tg(chat_id, "❌ Session expirée. Envoyez une nouvelle photo.")
                elif cdata == 'xl':
                    do_excel(chat_id, uid)
                elif cdata == 'pdf':
                    do_pdf(chat_id, uid)
                else:
                    do_excel(chat_id, uid)
                    do_pdf(chat_id, uid)
            elif cdata == 'email_send':
                executor.submit(do_send_email, chat_id, uid)
            elif cdata == 'project_create':
                executor.submit(do_create_project, chat_id, uid)
            elif cdata.startswith('ods_status:'):
                executor.submit(do_update_ods_status, chat_id, uid, cdata.split(':', 1)[1])
            elif cdata.startswith('offer_status:'):
                _, status, ref = cdata.split(':', 2)
                executor.submit(do_update_ods_status, chat_id, uid, status, ref)
            elif cdata == 'invoice_start':
                show_invoice_options(chat_id, uid)
            elif cdata.startswith('invoice_pct:'):
                show_invoice_preview(
                    chat_id, uid,
                    percentage=float(cdata.split(':', 1)[1]),
                )
            elif cdata == 'invoice_pct_other':
                d = user_data.get(uid, {})
                d['waiting_invoice_percentage'] = True
                user_data[uid] = d
                save_user_data()
                tg(chat_id, "📊 Entrez le pourcentage à facturer (ex. 40) :")
            elif cdata == 'invoice_fixed':
                d = user_data.get(uid, {})
                d['waiting_invoice_amount'] = True
                user_data[uid] = d
                save_user_data()
                tg(chat_id, "💵 Entrez le montant avant taxes (ex. 1250) :")
            elif cdata == 'invoice_confirm':
                executor.submit(do_issue_invoice, chat_id, uid)
            elif cdata == 'invoice_cancel':
                d = user_data.get(uid, {})
                d['pending_invoice'] = None
                d['waiting_invoice_percentage'] = False
                d['waiting_invoice_amount'] = False
                user_data[uid] = d
                save_user_data()
                tg(chat_id, "✅ Facture non émise. Vous pourrez la créer plus tard.")
            elif cdata.startswith('offer_pick:'):
                show_offer_conversion_confirmation(chat_id, uid, cdata.split(':', 1)[1])
            elif cdata.startswith('offer_convert:'):
                executor.submit(do_create_project, chat_id, uid, cdata.split(':', 1)[1])
            elif cdata == 'offer_search':
                d = user_data.get(uid, {})
                d['waiting_offer_search'] = True
                user_data[uid] = d
                save_user_data()
                tg(chat_id, "🔎 Entrez le numéro ODS (ex. ODS26-096 ou 096) :")
            elif cdata == 'email_cancel':
                tg(chat_id, "✅ Courriel non envoyé. Le PDF reste disponible dans Telegram.")
            elif cdata == 'num_ok':
                d = user_data.get(uid, {})
                d['project_num'] = d.get('suggested_num', '000')
                d['waiting_field'] = None
                user_data[uid] = d
                save_user_data()
                ask_next_missing(chat_id, uid)
            elif cdata == 'num_edit':
                d = user_data.get(uid, {})
                d['waiting_field'] = 'project_num'
                user_data[uid] = d
                save_user_data()
                tg(chat_id, "🔢 Entrez le numéro (ex: 082):")
            elif cdata == 'price_ok':
                d = user_data.get(uid, {})
                d['price_confirmed'] = True
                d['waiting_field'] = None
                user_data[uid] = d
                save_user_data()
                ask_next_missing(chat_id, uid)
            elif cdata == 'addr_ok':
                d = user_data.get(uid, {})
                d['addr_confirmed'] = True
                d['waiting_field'] = None
                user_data[uid] = d
                save_user_data()
                ask_next_missing(chat_id, uid)
            elif cdata == 'addr_edit':
                d = user_data.get(uid, {})
                d['addr_confirmed'] = False
                d['waiting_field'] = 'addr'
                user_data[uid] = d
                save_user_data()
                tg(chat_id, "📍 Entrez l'adresse complète:\n(ex: 247 Rue Beaumont\nGranby QC J2G 8S4\nCanada)")
            elif cdata.startswith('desc_'):
                d = user_data.get(uid, {})
                if cdata == 'desc_custom':
                    d['waiting_field'] = 'desc'
                    d['desc_confirmed'] = False
                    user_data[uid] = d
                    save_user_data()
                    tg(
                        chat_id,
                        "✏️ Collez votre contenu corrigé.\n\n"
                        "Vous pouvez envoyer :\n"
                        "• uniquement le nouveau mandat;\n"
                        "• uniquement les lignes de services;\n"
                        "• ou les deux avec les titres « Mandat : » et « Services : »."
                    )
                else:
                    idx = int(cdata.split('_')[1])
                    options = d.get('desc_options', [])
                    if idx < len(options):
                        d['desc'] = options[idx]
                        d['desc_confirmed'] = True
                        d['waiting_field'] = None
                        user_data[uid] = d
                        save_user_data()
                        tg(chat_id, "✅ Description sélectionnée.")
                        ask_next_missing(chat_id, uid)
            elif cdata == 'nouveau':
                user_data.pop(uid, None)
                save_user_data()
                tg(chat_id, "👋 *Nouveau client*\n\n📸 Envoyez une photo ou collez le texte du client.")
            elif cdata == 'price':
                d = user_data.get(uid, {})
                d['waiting_price'] = True
                d['waiting_field'] = None
                user_data[uid] = d
                save_user_data()
                tg(chat_id, "💰 Entrez le nouveau prix (ex: 3500):")
            elif cdata in ('tax_extra', 'tax_included'):
                d = user_data.get(uid, {})
                d['taxes'] = 'En sus' if cdata == 'tax_extra' else 'Incluses'
                user_data[uid] = d
                save_user_data()
                ask_next_missing(chat_id, uid)
            elif cdata == 'note_none':
                d = user_data.get(uid, {})
                d['special_note'] = ''
                d['special_note_confirmed'] = True
                user_data[uid] = d
                save_user_data()
                ask_next_missing(chat_id, uid)
            elif cdata == 'note_add':
                d = user_data.get(uid, {})
                d['waiting_field'] = 'special_note'
                user_data[uid] = d
                save_user_data()
                tg(chat_id, MISSING_QUESTIONS['special_note'])
            elif cdata in ('assumption_access_yes', 'assumption_access_no'):
                d = user_data.get(uid, {})
                d['assumption_access'] = cdata.endswith('_yes')
                d['assumption_access_confirmed'] = True
                user_data[uid] = d
                save_user_data()
                ask_next_missing(chat_id, uid)
            elif cdata in ('assumption_architect_yes', 'assumption_architect_no'):
                d = user_data.get(uid, {})
                d['assumption_architect'] = cdata.endswith('_yes')
                d['assumption_architect_confirmed'] = True
                user_data[uid] = d
                save_user_data()
                ask_next_missing(chat_id, uid)

        elif msg:
            uid = str(msg['from']['id'])
            chat_id = msg['chat']['id']
            if msg.get('text'):
                text = msg['text']
                if text in ('/start', '/nouveau', '📝 Nouvelle offre'):
                    user_data.pop(uid, None)
                    save_user_data()
                    tg(
                        chat_id,
                        f"👋 {DISPLAY_BRAND_SHORT} — Nouvelle offre\n\n"
                        "Envoyez une photo ou collez le texte/courriel du client.",
                        reply_markup=main_menu(),
                    )
                    return
                if text in ('/annuler', '/cancel', '❌ Annuler'):
                    user_data.pop(uid, None)
                    save_user_data()
                    tg(chat_id, "✅ Opération annulée.", reply_markup=main_menu())
                    return
                if text in ('/aide', '/help', '❓ Aide'):
                    tg(
                        chat_id,
                        "1. Envoyez une photo ou le texte du client.\n"
                        "2. Vérifiez les informations extraites.\n"
                        "3. Choisissez la description et le prix.\n"
                        "4. Générez le PDF ou Excel.",
                        reply_markup=main_menu(),
                    )
                    return
                if text == '📷 Envoyer une photo':
                    tg(chat_id, "📷 Envoyez maintenant une photo claire du message ou document du client.", reply_markup=main_menu())
                    return
                if text == '📋 Coller un texte':
                    tg(chat_id, "📋 Collez ici le courriel ou le texte complet du client.", reply_markup=main_menu())
                    return
                if text == '📁 Convertir une offre en projet':
                    show_pending_offers(chat_id, uid)
                    return
                d = user_data.get(uid, {})
                if d.get('waiting_offer_search'):
                    d['waiting_offer_search'] = False
                    user_data[uid] = d
                    save_user_data()
                    show_pending_offers(chat_id, uid, text)
                elif d.get('waiting_invoice_percentage'):
                    try:
                        percentage = float(text.strip().replace('%', '').replace(',', '.'))
                        if percentage <= 0 or percentage > 100:
                            raise ValueError
                        d['waiting_invoice_percentage'] = False
                        user_data[uid] = d
                        save_user_data()
                        show_invoice_preview(chat_id, uid, percentage=percentage)
                    except Exception:
                        tg(chat_id, "❌ Pourcentage invalide (ex. 25 ou 40).")
                elif d.get('waiting_invoice_amount'):
                    try:
                        amount = float(
                            text.strip().replace('$', '').replace(' ', '').replace(',', '')
                        )
                        if amount <= 0:
                            raise ValueError
                        d['waiting_invoice_amount'] = False
                        user_data[uid] = d
                        save_user_data()
                        show_invoice_preview(chat_id, uid, fixed_amount=amount)
                    except Exception:
                        tg(chat_id, "❌ Montant invalide (ex. 1250).")
                elif d.get('waiting_price'):
                    try:
                        price = int(text.strip().replace('$','').replace(',','').replace(' ',''))
                        d['price'] = price
                        d['waiting_price'] = False
                        d['price_confirmed'] = True
                        d['waiting_field'] = None
                        user_data[uid] = d
                        save_user_data()
                        ask_next_missing(chat_id, uid)
                    except:
                        tg(chat_id, "❌ Nombre invalide (ex: 3500)")
                elif d.get('waiting_field'):
                    field = d['waiting_field']
                    if False:
                        pass
                    else:
                        if field == 'desc':
                            mandate, service_lines = parse_custom_technical_content(
                                text,
                                d.get('desc', ''),
                            )
                            d['desc'] = mandate
                            if service_lines:
                                d['service_lines'] = service_lines
                        else:
                            d[field] = text.strip()
                        d['waiting_field'] = None
                        if field == 'addr':
                            d['addr_confirmed'] = True
                        if field == 'desc':
                            d['desc_confirmed'] = True
                        if field == 'special_note':
                            if text.strip().lower() in ('aucune', 'none', 'non'):
                                d['special_note'] = ''
                            d['special_note_confirmed'] = True
                        user_data[uid] = d
                        save_user_data()
                        ask_next_missing(chat_id, uid)
                else:
                    if len(text) > 50:
                        tg(chat_id, "🔍 Extraction en cours...")
                        executor.submit(do_extract_text, chat_id, uid, text)
                    else:
                        tg(chat_id, "📸 Envoyez une photo ou collez le texte du client.")
            elif msg.get('photo'):
                file_id = msg['photo'][-1]['file_id']
                queue_photo_extraction(chat_id, uid, file_id)
    except Exception as e:
        import traceback
        logger.error(f"handle_update error: {e}\n{traceback.format_exc()}")

@app.route('/webhook/telegram', methods=['POST'])
def webhook():
    if WEBHOOK_SECRET:
        supplied = request.headers.get('X-Telegram-Bot-Api-Secret-Token', '')
        if not hmac.compare_digest(supplied, WEBHOOK_SECRET):
            return 'forbidden', 403
    data = request.get_json(force=True, silent=True)
    if data:
        executor.submit(handle_update, data)
    return 'ok', 200

@app.route('/setup')
def setup():
    """Register the secured Telegram webhook and command menu."""
    if not has_setup_access():
        return 'forbidden', 403
    railway_url = os.environ.get('PUBLIC_URL', '').rstrip('/')
    if not railway_url:
        domain = os.environ.get('RAILWAY_PUBLIC_DOMAIN', '')
        railway_url = f"https://{domain}" if domain else ''
    if not railway_url:
        return jsonify({'error': 'PUBLIC_URL is required'}), 400

    webhook_payload = {
        'url': f"{railway_url}/webhook/telegram",
        'allowed_updates': ['message', 'callback_query'],
    }
    if WEBHOOK_SECRET:
        webhook_payload['secret_token'] = WEBHOOK_SECRET

    webhook_result = req.post(
        f'https://api.telegram.org/bot{BOT_TOKEN}/setWebhook',
        json=webhook_payload,
        timeout=15,
    ).json()
    commands_result = req.post(
        f'https://api.telegram.org/bot{BOT_TOKEN}/setMyCommands',
        json={
            'commands': [
                {'command': 'start', 'description': 'Démarrer et afficher le menu'},
                {'command': 'nouveau', 'description': 'Créer une nouvelle offre'},
                {'command': 'aide', 'description': "Afficher le guide d'utilisation"},
                {'command': 'annuler', 'description': "Annuler l'opération en cours"},
            ]
        },
        timeout=15,
    ).json()
    return jsonify({
        'ok': bool(webhook_result.get('ok') and commands_result.get('ok')),
        'webhook': webhook_result,
        'commands': commands_result,
    })


@app.route('/status')
def status():
    if not has_setup_access():
        return jsonify({'error': 'forbidden'}), 403
    return jsonify({'status': 'ok', 'active_sessions': len(user_data)})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
