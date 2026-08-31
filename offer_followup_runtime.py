import io
import html
import logging
import os
import re
import threading
from datetime import datetime
from zoneinfo import ZoneInfo

import openpyxl
import requests

import menu_guard_runtime as guarded
from offer_followup import (
    FollowupStore, build_followup_email, followup_stage, is_open_offer,
    normalized_status, parse_date, recommended_followup_day,
)


app = guarded.app
legacy = guarded.legacy
logger = logging.getLogger(__name__)

STORE = FollowupStore(os.path.join(legacy.DATA_DIR, "offer_followups.json"))
TIMEZONE_NAME = os.environ.get("OFFER_FOLLOWUP_TIMEZONE", "America/Toronto")
_base_main_menu = guarded.final_main_menu


def local_now():
    try:
        return datetime.now(ZoneInfo(TIMEZONE_NAME))
    except Exception:
        return datetime.now(ZoneInfo("UTC"))


def followup_main_menu():
    menu = _base_main_menu()
    keyboard = list(menu.get("keyboard") or [])
    keyboard.insert(4 if len(keyboard) >= 4 else len(keyboard), [{"text": "📬 Suivi offres"}])
    menu["keyboard"] = keyboard
    return menu


guarded.final_main_menu = followup_main_menu
guarded.MAIN_MENU_COMMANDS.add("📬 Suivi offres")
legacy.main_menu = followup_main_menu


def _ack(callback):
    try:
        requests.post(
            f"https://api.telegram.org/bot{legacy.BOT_TOKEN}/answerCallbackQuery",
            json={"callback_query_id": callback.get("id")}, timeout=3,
        )
    except Exception:
        pass


def _reference(value):
    text = str(value or "").strip()
    match = re.search(r"ODS\d{2}-\d{3}(?:-[A-Z]{3})?", text, re.I)
    return match.group(0).upper() if match else text[:80]


def _amount(value):
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value or "0").replace("$", "").replace(" ", "")
    if "," in raw and "." in raw:
        raw = raw.replace(",", "")
    elif "," in raw:
        raw = raw.replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        return 0.0


def load_open_offers():
    config = legacy.microsoft_email_config()
    token = legacy.graph_access_token()
    content = legacy.download_onedrive_path(token, config["EMAIL_SENDER"], legacy.ODS_LIST_PATH)
    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    offers = []
    for ws in workbook.worksheets:
        if not str(ws.title).lower().startswith("data "):
            continue
        headers = {str(cell.value or "").strip(): cell.column for cell in ws[1] if cell.value is not None}
        required = ("Description", "Date", "Status", "Price($)")
        if any(name not in headers for name in required):
            continue
        for row in range(2, ws.max_row + 1):
            description = str(ws.cell(row, headers["Description"]).value or "").strip()
            if not description:
                continue
            raw_date = ws.cell(row, headers["Date"]).value
            offer = {
                "reference": _reference(description),
                "description": description,
                "date": raw_date.isoformat() if hasattr(raw_date, "isoformat") else str(raw_date or ""),
                "status": normalized_status(ws.cell(row, headers["Status"]).value),
                "price": _amount(ws.cell(row, headers["Price($)"]).value),
                "contact": str(ws.cell(row, headers.get("Contact", 0)).value or "").strip() if headers.get("Contact") else "",
                "email": str(ws.cell(row, headers.get("Email", 0)).value or "").strip() if headers.get("Email") else "",
                "source": str(ws.cell(row, headers.get("Source", 0)).value or "").strip() if headers.get("Source") else "",
                "sheet": ws.title,
                "row": row,
            }
            if is_open_offer(offer):
                offers.append(offer)
    offers.sort(key=lambda item: (followup_stage(item["date"], local_now().date())["days"], item["reference"]), reverse=True)
    return offers


def update_list_status(reference, status):
    if status not in {"Hold", "Refused", "Closed", "In process"}:
        raise ValueError("statut non valide")
    with legacy.ods_list_lock:
        config = legacy.microsoft_email_config()
        token = legacy.graph_access_token()
        owner = config["EMAIL_SENDER"]
        content = legacy.download_onedrive_path(token, owner, legacy.ODS_LIST_PATH)
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        found = False
        for ws in workbook.worksheets:
            headers = {str(cell.value or "").strip(): cell.column for cell in ws[1] if cell.value is not None}
            if "Description" not in headers or "Status" not in headers:
                continue
            for row in range(2, ws.max_row + 1):
                description = str(ws.cell(row, headers["Description"]).value or "")
                if reference.upper() in description.upper():
                    ws.cell(row, headers["Status"]).value = status
                    found = True
                    break
            if found:
                break
        if not found:
            raise RuntimeError("offre introuvable dans List.xlsx")
        output = io.BytesIO()
        workbook.save(output)
        legacy.upload_onedrive_path(
            token, owner, legacy.ODS_LIST_PATH, output.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def update_list_email(reference, email_address):
    with legacy.ods_list_lock:
        config = legacy.microsoft_email_config()
        token = legacy.graph_access_token()
        owner = config["EMAIL_SENDER"]
        content = legacy.download_onedrive_path(token, owner, legacy.ODS_LIST_PATH)
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        found = False
        for ws in workbook.worksheets:
            headers = {str(cell.value or "").strip(): cell.column for cell in ws[1] if cell.value is not None}
            if "Description" not in headers or "Email" not in headers:
                continue
            for row in range(2, ws.max_row + 1):
                description = str(ws.cell(row, headers["Description"]).value or "")
                if reference.upper() in description.upper():
                    ws.cell(row, headers["Email"]).value = email_address
                    found = True
                    break
            if found:
                break
        if not found:
            raise RuntimeError("offre introuvable dans List.xlsx")
        output = io.BytesIO()
        workbook.save(output)
        legacy.upload_onedrive_path(
            token, owner, legacy.ODS_LIST_PATH, output.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def _email_from_text(value):
    match = re.search(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", str(value or ""), re.I)
    return legacy.valid_client_email(match.group(0)) if match else ""


def _history_email(reference):
    for records in (legacy.offers_history or {}).values():
        for ref, record in (records or {}).items():
            data = (record or {}).get("data") or {}
            if reference.upper() in str(ref).upper() or reference.upper() in str(data.get("odsNum") or "").upper():
                email_address = legacy.valid_client_email(data.get("email"))
                if email_address:
                    return email_address
    return ""


def _archive_email(reference, token, owner, archive_items):
    candidates = [
        str(item.get("name") or "") for item in archive_items
        if str(item.get("name") or "").lower().endswith(".xlsx")
        and reference.upper() in str(item.get("name") or "").upper()
    ]
    for filename in candidates:
        try:
            content = legacy.download_onedrive_path(
                token, owner, f"Metra Structure Inc/Offre de service/{filename}"
            )
            workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = workbook["ODS"] if "ODS" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
            # The approved ODS template stores the client email in B10. Search
            # only the client-information block as a fallback so the company
            # signature email is never mistaken for the client's address.
            for value in [ws["B10"].value] + [ws.cell(row, 2).value for row in range(7, 16)]:
                email_address = _email_from_text(value)
                if email_address and email_address.lower() not in {
                    str(owner).lower(), "accounting@metrastructure.ca"
                }:
                    return email_address
        except Exception as exc:
            logger.warning("August email recovery skipped %s: %s", filename, exc)
    return ""


def complete_august_2026_emails():
    """Fill only missing client emails in August 2026 rows of List.xlsx."""
    with legacy.ods_list_lock:
        config = legacy.microsoft_email_config()
        token = legacy.graph_access_token()
        owner = config["EMAIL_SENDER"]
        content = legacy.download_onedrive_path(token, owner, legacy.ODS_LIST_PATH)
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        if "data 2026" not in workbook.sheetnames:
            raise RuntimeError("onglet data 2026 introuvable")
        ws = workbook["data 2026"]
        headers = {str(cell.value or "").strip(): cell.column for cell in ws[1] if cell.value is not None}
        for required in ("Description", "Date", "Email"):
            if required not in headers:
                raise RuntimeError(f"colonne {required} introuvable")
        archive_items = legacy.list_onedrive_children(
            token, owner, "Metra Structure Inc/Offre de service"
        )
        stats = {"august_rows": 0, "already_valid": 0, "filled": 0, "unresolved": []}
        for row in range(2, ws.max_row + 1):
            sent_at = parse_date(ws.cell(row, headers["Date"]).value)
            if not sent_at or sent_at.year != 2026 or sent_at.month != 8:
                continue
            description = str(ws.cell(row, headers["Description"]).value or "").strip()
            if not description:
                continue
            stats["august_rows"] += 1
            current = legacy.valid_client_email(ws.cell(row, headers["Email"]).value)
            if current:
                stats["already_valid"] += 1
                continue
            reference = _reference(description)
            email_address = _history_email(reference) or _archive_email(
                reference, token, owner, archive_items
            )
            if email_address:
                ws.cell(row, headers["Email"]).value = email_address
                stats["filled"] += 1
            else:
                stats["unresolved"].append(reference)
        if stats["filled"]:
            output = io.BytesIO()
            workbook.save(output)
            legacy.upload_onedrive_path(
                token, owner, legacy.ODS_LIST_PATH, output.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        return stats


def _run_august_email_backfill():
    # Run once on the active production service only. The marker lives on the
    # persistent Railway volume and prevents repeated writes after restarts.
    marker = "august_email_backfill_2026_v1"
    if STORE.scheduler_value(marker):
        return
    threading.Event().wait(20)
    try:
        stats = complete_august_2026_emails()
        STORE.scheduler_value(marker, local_now().isoformat())
        unresolved = stats["unresolved"]
        message = (
            "📧 Complétion des courriels — août 2026\n\n"
            f"Lignes d’août : {stats['august_rows']}\n"
            f"Déjà valides : {stats['already_valid']}\n"
            f"Courriels ajoutés : {stats['filled']}\n"
            f"Introuvables : {len(unresolved)}"
        )
        if unresolved:
            message += "\n\nÀ compléter manuellement :\n" + "\n".join(f"• {ref}" for ref in unresolved[:25])
        for chat_id in sorted(legacy.ALLOWED_USERS):
            legacy.tg(chat_id, message)
    except Exception as exc:
        logger.exception("August 2026 email backfill failed")
        for chat_id in sorted(legacy.ALLOWED_USERS):
            legacy.tg(chat_id, f"❌ Complétion des courriels d’août impossible : {exc}")


def _session(uid):
    return legacy.user_data.setdefault(str(uid), {})


def _save_session(uid, session):
    legacy.user_data[str(uid)] = session
    legacy.save_user_data()


def _month_key(value):
    parsed = parse_date(value)
    return parsed.strftime("%Y-%m") if parsed else "unknown"


def show_offer_months(chat_id, uid):
    try:
        offers = load_open_offers()
    except Exception as exc:
        logger.exception("Open-offer follow-up list failed")
        legacy.tg(chat_id, f"❌ Lecture de List.xlsx impossible : {exc}")
        return
    months = {}
    for offer in offers:
        months.setdefault(_month_key(offer.get("date")), []).append(offer)
    session = _session(uid)
    session["offer_followup_months"] = months
    session.pop("offer_followup_month", None)
    session.pop("offer_followup_selected", None)
    _save_session(uid, session)
    rows = []
    month_names = {
        1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril",
        5: "Mai", 6: "Juin", 7: "Juillet", 8: "Août",
        9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Décembre",
    }
    for key in sorted(months, reverse=True):
        if key == "unknown":
            label = "Date inconnue"
        else:
            year, month = key.split("-")
            label = f"{month_names.get(int(month), month)} {year}"
        rows.append([{"text": f"📅 {label} — {len(months[key])} offre(s)", "callback_data": f"of_month:{key}"}])
    if not rows:
        legacy.tg(chat_id, "✅ Aucune offre ouverte dans List.xlsx.", reply_markup=followup_main_menu())
        return
    legacy.tg(chat_id, "📬 Suivi des offres ouvertes\n\nChoisissez le mois :", rows)


def show_open_offers(chat_id, uid, month_key):
    session = _session(uid)
    offers = (session.get("offer_followup_months") or {}).get(month_key)
    if offers is None:
        show_offer_months(chat_id, uid)
        return
    choices = {}
    rows = []
    for index, offer in enumerate(offers[:40], start=1):
        choices[str(index)] = offer
        stage = followup_stage(offer["date"], local_now().date())
        warning = "🔴 " if stage["days"] >= 60 else "🟠 " if stage["days"] >= 30 else "⚪ "
        label = f"{warning}{stage['days']}j · {offer['reference']} · {offer['contact'] or 'Client'}"
        rows.append([{"text": label[:62], "callback_data": f"of_pick:{index}"}])
    session["offer_followup_choices"] = choices
    session["offer_followup_month"] = month_key
    session.pop("offer_followup_selected", None)
    _save_session(uid, session)
    rows.append([{"text": "⬅️ Choisir un autre mois", "callback_data": "of_months"}])
    legacy.tg(chat_id, f"📬 Offres du mois {month_key}\n\n{len(offers)} offre(s). Choisissez une offre :", rows)


def _selected(uid):
    session = _session(uid)
    return session.get("offer_followup_selected")


def _offer_text(offer, state):
    stage = followup_stage(offer.get("date"), local_now().date())
    sent = offer.get("date")
    sent_text = sent.strftime("%Y-%m-%d") if hasattr(sent, "strftime") else str(sent or "—")[:10]
    last = str(state.get("last_followup_at") or "Jamais")
    since_last = "—"
    last_date = parse_date(state.get("last_followup_at"))
    if last_date:
        since_last = f"{(local_now().date() - last_date).days} jour(s)"
    return (
        f"📬 {offer['reference']}\n\n"
        f"Client : {offer.get('contact') or '—'}\n"
        f"Courriel : {offer.get('email') or '—'}\n"
        f"Montant : {offer.get('price', 0):,.0f} $\n"
        f"Envoyée : {sent_text} ({stage['days']} jours)\n"
        f"Statut : {offer.get('status')}\n"
        f"Âge de l’offre : {stage['days']} jour(s)\n"
        f"Relances enregistrées : {state.get('followup_count', 0)}\n"
        f"Dernière relance : {last}\n"
        f"Depuis la dernière relance : {since_last}"
    )


def show_offer(chat_id, uid):
    offer = _selected(uid)
    if not offer:
        legacy.tg(chat_id, "❌ Offre introuvable. Ouvrez de nouveau Suivi offres.")
        return
    state = STORE.load().get("offers", {}).get(offer["reference"], {})
    buttons = [
        [{"text": "✉️ Préparer un courriel de suivi", "callback_data": "of_email_menu"}],
        [{"text": "✅ Relance effectuée", "callback_data": "of_mark"}],
        [
            {"text": "⏸ Hold", "callback_data": "of_status:Hold"},
            {"text": "🔄 In process", "callback_data": "of_status:In process"},
        ],
        [
            {"text": "❌ Refused", "callback_data": "of_status:Refused"},
            {"text": "🔒 Closed", "callback_data": "of_status:Closed"},
        ],
        [{"text": "📁 Acceptée → projet", "callback_data": "of_convert"}],
        [{"text": "⬅️ Offres de ce mois", "callback_data": "of_back"}],
    ]
    legacy.tg(chat_id, _offer_text(offer, state), buttons)


def show_email_menu(chat_id, uid):
    offer = _selected(uid)
    if not offer:
        legacy.tg(chat_id, "❌ Offre introuvable.")
        return
    age = followup_stage(offer.get("date"), local_now().date())["days"]
    recommended = recommended_followup_day(age)
    buttons = []
    for left, right in ((3, 7), (10, 15)):
        row = []
        for day in (left, right):
            star = " ⭐" if day == recommended else ""
            row.append({"text": f"Suivi {day} jours{star}", "callback_data": f"of_email:{day}"})
        buttons.append(row)
    star = " ⭐" if recommended == 30 else ""
    buttons.append([{"text": f"Suivi final 30 jours{star}", "callback_data": "of_email:30"}])
    buttons.append([{"text": "⬅️ Retour", "callback_data": "of_detail"}])
    legacy.tg(
        chat_id,
        f"✉️ Courriel de suivi\n\nCette offre a été envoyée il y a {age} jour(s).\n"
        f"Étape suggérée : suivi {recommended} jours.\n\nChoisissez le modèle :",
        buttons,
    )


def show_email_preview(chat_id, uid, day):
    offer = _selected(uid)
    if not offer:
        legacy.tg(chat_id, "❌ Offre introuvable.")
        return
    recipient = legacy.valid_client_email(offer.get("email"))
    if not recipient:
        for _ref, record in (legacy.offers_history.get(str(uid), {}) or {}).items():
            data = (record or {}).get("data") or {}
            if offer["reference"].upper() in str(_ref).upper() or offer["reference"].upper() in str(data.get("odsNum") or "").upper():
                recipient = legacy.valid_client_email(data.get("email"))
                if recipient:
                    offer["email"] = recipient
                    break
    try:
        subject, body = build_followup_email(offer, int(day))
    except (TypeError, ValueError):
        legacy.tg(chat_id, "❌ Modèle de suivi non valide.")
        return
    session = _session(uid)
    session["pending_offer_followup_email"] = {
        "reference": offer["reference"], "recipient": recipient,
        "subject": subject, "body": body, "day": int(day),
    }
    _save_session(uid, session)
    recipient_display = recipient or "⚠️ Courriel client manquant"
    buttons = [
        [{"text": "🧪 Envoyer un test à moi-même", "callback_data": "of_email_test"}],
    ]
    if recipient:
        buttons.append([{"text": "✅ Confirmer et envoyer au client", "callback_data": "of_email_send"}])
    else:
        buttons.append([{"text": "✏️ Ajouter le courriel client", "callback_data": "of_email_add"}])
    buttons.extend([
        [{"text": "✏️ Choisir un autre modèle", "callback_data": "of_email_menu"}],
        [{"text": "❌ Annuler", "callback_data": "of_email_cancel"}],
    ])
    legacy.tg(
        chat_id,
        f"✉️ Aperçu du courriel\n\nÀ : {recipient_display}\nObjet : {subject}\n\n{body}",
        buttons,
    )


def _followup_html(body):
    raw = str(body or "").strip()
    closing = "Bien cordialement,"
    message, _separator, _signature = raw.partition(closing)
    paragraphs = [part.strip() for part in message.split("\n\n") if part.strip()]
    rendered = "".join(
        f'<p style="margin:0 0 16px 0">{html.escape(part).replace(chr(10), "<br>")}</p>'
        for part in paragraphs
    )
    return (
        '<div style="font-family:Arial,Helvetica,sans-serif;font-size:11pt;line-height:1.55;color:#252525;max-width:680px">'
        f'{rendered}'
        '<p style="margin:22px 0 16px 0">Bien cordialement,</p>'
        '<div style="border-left:3px solid #f5a400;padding:2px 0 2px 12px">'
        '<div style="font-weight:700;color:#162b67">Arash Rohani, ing., P.Eng.</div>'
        '<div>Président – Ingénieur en structure et en génie civil</div>'
        '<div style="font-weight:700">Metra Consultation Inc.</div>'
        '<div style="margin-top:4px">'
        '<a href="mailto:arash.rohani@metrastructure.ca" style="color:#225ea8;text-decoration:none">arash.rohani@metrastructure.ca</a>'
        ' &nbsp;|&nbsp; '
        '<a href="tel:+14388674131" style="color:#225ea8;text-decoration:none">(438) 867-4131</a>'
        '</div></div></div>'
    )


def send_followup_email(payload):
    config = legacy.microsoft_email_config()
    token = legacy.graph_access_token()
    sender = requests.utils.quote(config["EMAIL_SENDER"], safe="")
    response = requests.post(
        f"https://graph.microsoft.com/v1.0/users/{sender}/sendMail",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        json={
            "message": {
                "subject": payload["subject"],
                "body": {"contentType": "HTML", "content": _followup_html(payload["body"])},
                "toRecipients": [{"emailAddress": {"address": payload["recipient"]}}],
            },
            "saveToSentItems": True,
        },
        timeout=45,
    )
    if response.status_code != 202:
        raise RuntimeError(f"Microsoft 365 a refusé l’envoi ({response.status_code})")


def do_send_followup_email(chat_id, uid, test_only=False):
    session = _session(uid)
    payload = dict(session.get("pending_offer_followup_email") or {})
    if not payload:
        legacy.tg(chat_id, "❌ Aperçu expiré. Préparez de nouveau le courriel.")
        return
    if not test_only and not legacy.valid_client_email(payload.get("recipient")):
        legacy.tg(chat_id, "⚠️ Ajoutez d’abord une adresse courriel client valide.")
        return
    if session.get("offer_followup_email_sending"):
        legacy.tg(chat_id, "⏳ Envoi déjà en cours.")
        return
    session["offer_followup_email_sending"] = True
    _save_session(uid, session)
    try:
        outgoing = dict(payload)
        if test_only:
            outgoing["recipient"] = legacy.microsoft_email_config()["EMAIL_SENDER"]
            outgoing["subject"] = "[TEST] " + outgoing["subject"]
        send_followup_email(outgoing)
        session = _session(uid)
        session["offer_followup_email_sending"] = False
        if not test_only:
            STORE.mark_followed(payload["reference"], local_now().date(), payload["day"])
            session.pop("pending_offer_followup_email", None)
        _save_session(uid, session)
        if test_only:
            legacy.tg(
                chat_id,
                f"🧪 Courriel test envoyé à {outgoing['recipient']}.\n"
                "Vérifiez-le, puis confirmez séparément l’envoi réel au client.",
                [
                    [{"text": "✅ Envoyer maintenant au client", "callback_data": "of_email_send"}],
                    [{"text": "✏️ Choisir un autre modèle", "callback_data": "of_email_menu"}],
                    [{"text": "❌ Annuler", "callback_data": "of_email_cancel"}],
                ],
            )
            return
        message = (
            f"✅ Courriel de suivi {payload['day']} jours envoyé à {payload['recipient']}."
        )
        if int(payload["day"]) == 30:
            message += "\n\nSans réponse, vous pouvez ensuite choisir « Closed » dans la fiche de l’offre."
        legacy.tg(chat_id, message)
        show_offer(chat_id, uid)
    except Exception as exc:
        session = _session(uid)
        session["offer_followup_email_sending"] = False
        _save_session(uid, session)
        logger.exception("Offer follow-up email failed")
        legacy.tg(chat_id, f"❌ Envoi impossible : {exc}")


def _history_reference(uid, reference):
    for ref, record in (legacy.offers_history.get(str(uid), {}) or {}).items():
        data = (record or {}).get("data") or {}
        if reference.upper() in str(ref).upper() or reference.upper() in str(data.get("odsNum") or "").upper():
            return ref
    return ""


_previous_handle_update = legacy.handle_update


def handle_update_offer_followup(data):
    msg = data.get("message") or {}
    callback = data.get("callback_query") or {}
    actor_id = (callback.get("from") or msg.get("from") or {}).get("id")
    chat_id = ((callback.get("message") or {}).get("chat") or {}).get("id") if callback else (msg.get("chat") or {}).get("id")
    if msg and actor_id in legacy.ALLOWED_USERS:
        uid = str(actor_id)
        session = _session(uid)
        if session.get("waiting_offer_followup_email") and msg.get("text"):
            email_address = legacy.valid_client_email(str(msg.get("text") or "").strip())
            if not email_address:
                legacy.tg(chat_id, "❌ Adresse invalide. Exemple : client@example.com")
                return
            offer = _selected(uid)
            pending = session.get("pending_offer_followup_email") or {}
            try:
                update_list_email(offer["reference"], email_address)
                offer["email"] = email_address
                pending["recipient"] = email_address
                session["offer_followup_selected"] = offer
                session["pending_offer_followup_email"] = pending
                session["waiting_offer_followup_email"] = False
                _save_session(uid, session)
                legacy.tg(chat_id, "✅ Courriel ajouté dans List.xlsx.")
                show_email_preview(chat_id, uid, pending.get("day"))
            except Exception as exc:
                legacy.tg(chat_id, f"❌ Enregistrement impossible : {exc}")
            return
    if msg and str(msg.get("text") or "").strip() == "📬 Suivi offres":
        if actor_id not in legacy.ALLOWED_USERS:
            legacy.tg(chat_id, "⛔ Ce bot est privé.")
            return
        legacy.executor.submit(show_offer_months, chat_id, str(actor_id))
        return
    cdata = str(callback.get("data") or "")
    if cdata.startswith("of_"):
        if actor_id not in legacy.ALLOWED_USERS:
            legacy.tg(chat_id, "⛔ Ce bot est privé.")
            return
        _ack(callback)
        uid = str(actor_id)
        if cdata.startswith("of_month:"):
            month_key = cdata.split(":", 1)[1]
            show_open_offers(chat_id, uid, month_key)
        elif cdata == "of_months":
            legacy.executor.submit(show_offer_months, chat_id, uid)
        elif cdata.startswith("of_pick:"):
            choice = cdata.split(":", 1)[1]
            session = _session(uid)
            offer = (session.get("offer_followup_choices") or {}).get(choice)
            if not offer:
                legacy.tg(chat_id, "❌ Liste expirée. Ouvrez de nouveau Suivi offres.")
                return
            session["offer_followup_selected"] = offer
            _save_session(uid, session)
            show_offer(chat_id, uid)
        elif cdata == "of_mark":
            offer = _selected(uid)
            if not offer:
                legacy.tg(chat_id, "❌ Offre introuvable.")
                return
            STORE.mark_followed(offer["reference"], local_now().date())
            legacy.tg(chat_id, "✅ Relance enregistrée. Vous choisissez vous-même la prochaine action.")
            show_offer(chat_id, uid)
        elif cdata == "of_email_menu":
            show_email_menu(chat_id, uid)
        elif cdata.startswith("of_email:"):
            show_email_preview(chat_id, uid, cdata.split(":", 1)[1])
        elif cdata == "of_email_send":
            legacy.executor.submit(do_send_followup_email, chat_id, uid)
        elif cdata == "of_email_test":
            legacy.executor.submit(do_send_followup_email, chat_id, uid, True)
        elif cdata == "of_email_add":
            session = _session(uid)
            session["waiting_offer_followup_email"] = True
            _save_session(uid, session)
            legacy.tg(chat_id, "✏️ Entrez l’adresse courriel du client :")
        elif cdata == "of_email_cancel":
            session = _session(uid)
            session.pop("pending_offer_followup_email", None)
            _save_session(uid, session)
            legacy.tg(chat_id, "✅ Courriel non envoyé.")
            show_offer(chat_id, uid)
        elif cdata == "of_detail":
            show_offer(chat_id, uid)
        elif cdata.startswith("of_status:"):
            offer = _selected(uid)
            status = cdata.split(":", 1)[1]
            if not offer:
                legacy.tg(chat_id, "❌ Offre introuvable.")
                return
            try:
                update_list_status(offer["reference"], status)
                offer["status"] = status
                session = _session(uid)
                session["offer_followup_selected"] = offer
                _save_session(uid, session)
                legacy.tg(chat_id, f"✅ List.xlsx mis à jour : {status}")
                show_offer(chat_id, uid)
            except Exception as exc:
                legacy.tg(chat_id, f"❌ Mise à jour impossible : {exc}")
        elif cdata == "of_convert":
            offer = _selected(uid)
            ref = _history_reference(uid, offer["reference"] if offer else "")
            if not ref:
                legacy.tg(chat_id, "⚠️ Fichier de session introuvable. Utilisez « Convertir une offre en projet » pour rechercher cette ODS.")
                return
            legacy.show_offer_conversion_confirmation(chat_id, uid, ref)
        elif cdata == "of_back":
            month_key = str(_session(uid).get("offer_followup_month") or "")
            show_open_offers(chat_id, uid, month_key)
        return
    return _previous_handle_update(data)


legacy.handle_update = handle_update_offer_followup
logger.info("OPEN OFFER FOLLOW-UP RUNTIME ACTIVE")

if os.environ.get("RAILWAY_SERVICE_NAME", "") == "metra-offre-invoice-bot":
    threading.Thread(target=_run_august_email_backfill, name="august-email-backfill", daemon=True).start()
