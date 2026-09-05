import io
import logging
from datetime import datetime

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import openpyxl
import requests

import invoice_control_runtime as control

app = control.app
legacy = control.legacy
logger = logging.getLogger(__name__)


def _amount(value):
    if value in (None, ''):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).strip().replace('$', '').replace(' ', '')
    if ',' in cleaned and '.' in cleaned:
        cleaned = cleaned.replace(',', '')
    elif ',' in cleaned:
        cleaned = cleaned.replace(',', '.')
    try:
        return float(cleaned)
    except (TypeError, ValueError):
        return 0.0


def _date(value):
    if isinstance(value, datetime):
        return value
    if hasattr(value, 'year') and hasattr(value, 'month'):
        return datetime(value.year, value.month, getattr(value, 'day', 1))
    text = str(value or '').strip()
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(text[:10], fmt)
        except ValueError:
            continue
    return None


def _status(value):
    raw = str(value or 'In process').strip().lower()
    aliases = {
        'accept': 'Accept', 'accepted': 'Accept', 'accepté': 'Accept',
        'in process': 'In process', 'in progress': 'In process',
        'hold': 'Hold', 'on hold': 'Hold',
        'refused': 'Refused', 'rejected': 'Refused',
        'closed': 'Closed', 'close': 'Closed',
    }
    return aliases.get(raw, str(value or 'In process').strip().title())


MONTH_NAMES_FR = {
    1: 'Janvier', 2: 'Février', 3: 'Mars', 4: 'Avril', 5: 'Mai',
    6: 'Juin', 7: 'Juillet', 8: 'Août', 9: 'Septembre',
    10: 'Octobre', 11: 'Novembre', 12: 'Décembre',
}


def _period_bounds(period, now):
    if period == 'previous':
        end = datetime(now.year, now.month, 1)
        start = datetime(now.year - 1, 12, 1) if now.month == 1 else datetime(now.year, now.month - 1, 1)
    elif period == 'year':
        start, end = datetime(now.year, 1, 1), datetime(now.year + 1, 1, 1)
    elif len(str(period)) == 7 and str(period)[4] == '-':
        try:
            year, month = (int(part) for part in str(period).split('-', 1))
            start = datetime(year, month, 1)
            end = datetime(year + 1, 1, 1) if month == 12 else datetime(year, month + 1, 1)
        except (TypeError, ValueError):
            raise ValueError(f'période invalide : {period}')
    else:
        start = datetime(now.year, now.month, 1)
        end = datetime(now.year + 1, 1, 1) if now.month == 12 else datetime(now.year, now.month + 1, 1)
    label = str(start.year) if end.month == 1 and end.year == start.year + 1 and start.month == 1 else f'{MONTH_NAMES_FR[start.month]} {start.year}'
    return start, end, label


def _sheet_headers(ws):
    return {
        str(cell.value or '').strip(): cell.column
        for cell in ws[1] if cell.value is not None
    }


def available_months(workbook):
    """Return every offer month present in List.xlsx, newest first."""
    months = set()
    for ws in workbook.worksheets:
        if not ws.title.lower().startswith('data '):
            continue
        headers = _sheet_headers(ws)
        if 'Date' not in headers or 'Description' not in headers:
            continue
        for row in range(2, ws.max_row + 1):
            description = str(ws.cell(row, headers['Description']).value or '').strip()
            sent_at = _date(ws.cell(row, headers['Date']).value)
            if description and sent_at:
                months.add(f'{sent_at.year:04d}-{sent_at.month:02d}')
    return sorted(months, reverse=True)


def dashboard_metrics(workbook, period='month', marketing=False, now=None):
    """Aggregate management and source-attribution KPIs from List.xlsx."""
    now = now or datetime.now()
    start, end, label = _period_bounds(period, now)
    sheet_name = f'data {start.year}'
    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(f'onglet {sheet_name} introuvable dans List.xlsx')
    ws = workbook[sheet_name]
    headers = _sheet_headers(ws)
    required = ('Description', 'Price($)', 'Date', 'Status', 'Accepted Price', 'Source')
    missing = [name for name in required if name not in headers]
    if missing:
        raise RuntimeError('colonnes manquantes dans List.xlsx : ' + ', '.join(missing))

    rows = []
    for row in range(2, ws.max_row + 1):
        description = str(ws.cell(row, headers['Description']).value or '').strip()
        sent_at = _date(ws.cell(row, headers['Date']).value)
        if not description or not sent_at or not (start <= sent_at < end):
            continue
        status = _status(ws.cell(row, headers['Status']).value)
        price = _amount(ws.cell(row, headers['Price($)']).value)
        accepted_price = _amount(ws.cell(row, headers['Accepted Price']).value)
        if status == 'Accept' and accepted_price <= 0:
            accepted_price = price
        rows.append({
            'date': sent_at, 'status': status, 'price': price,
            'accepted_price': accepted_price,
            'source': str(ws.cell(row, headers['Source']).value or 'Non indiqué').strip() or 'Non indiqué',
        })

    statuses = {name: 0 for name in ('Accept', 'In process', 'Hold', 'Refused', 'Closed')}
    sources = {}
    for item in rows:
        statuses[item['status']] = statuses.get(item['status'], 0) + 1
        source = sources.setdefault(item['source'], {
            'offers': 0, 'accepted': 0, 'quoted': 0.0, 'revenue': 0.0,
        })
        source['offers'] += 1
        source['quoted'] += item['price']
        if item['status'] == 'Accept':
            source['accepted'] += 1
            source['revenue'] += item['accepted_price']

    total = len(rows)
    accepted = statuses.get('Accept', 0)
    stale = [item for item in rows if item['status'] == 'In process' and (now - item['date']).days >= 14]
    return {
        'period_label': label, 'marketing': marketing, 'offers': total,
        'quoted': sum(item['price'] for item in rows), 'accepted': accepted,
        'revenue': sum(item['accepted_price'] for item in rows if item['status'] == 'Accept'),
        'conversion': accepted / total if total else 0.0, 'statuses': statuses,
        'stale_count': len(stale), 'stale_value': sum(item['price'] for item in stale),
        'sources': sorted(
            sources.items(),
            key=lambda entry: (entry[1]['revenue'], entry[1]['accepted'], entry[1]['offers']),
            reverse=True,
        ),
    }


def _money(value):
    return f"{value:,.0f} $".replace(',', ' ')


def format_dashboard(metrics):
    title = '📣 Marketing' if metrics['marketing'] else '📊 Tableau de bord'
    text = [
        f"{title} — {metrics['period_label']}", '',
        f"📝 Offres : {metrics['offers']}  |  {_money(metrics['quoted'])}",
        f"✅ Acceptées : {metrics['accepted']}  |  {_money(metrics['revenue'])}",
        f"🎯 Taux de conversion : {metrics['conversion']:.1%}",
        f"🔄 Pipeline : {metrics['statuses'].get('In process', 0)} en cours  |  {metrics['statuses'].get('Hold', 0)} en attente",
        f"❌ Perdues : {metrics['statuses'].get('Refused', 0)} refusées  |  {metrics['statuses'].get('Closed', 0)} fermées",
        f"⏰ Relances +14 jours : {metrics['stale_count']}  |  {_money(metrics['stale_value'])}",
    ]
    if metrics['sources']:
        text.extend(['', '📣 Performance par source :'])
        for name, source in metrics['sources'][:10 if metrics['marketing'] else 5]:
            rate = source['accepted'] / source['offers'] if source['offers'] else 0
            text.append(
                f"• {name}: {source['offers']} offres | {source['accepted']} acceptées"
                f" | {rate:.0%} | {_money(source['revenue'])}"
            )
    else:
        text.extend(['', 'Aucune offre pour cette période.'])
    if metrics['marketing']:
        text.extend(['', 'ℹ️ Attribution basée sur la colonne Source de List.xlsx.'])
    return '\n'.join(text)


def dashboard_buttons(marketing=False, period='month'):
    view = 'marketing' if marketing else 'summary'
    alternate_view = 'summary' if marketing else 'marketing'
    alternate_label = '📊 Synthèse' if marketing else '📣 Marketing'
    return [
        [
            {'text': '📅 Ce mois', 'callback_data': f'ods_dashboard:{view}:month'},
            {'text': '◀️ Mois précédent', 'callback_data': f'ods_dashboard:{view}:previous'},
        ],
        [
            {'text': '🗓 Choisir un mois', 'callback_data': f'ods_dashboard_months:{view}:0'},
            {'text': '🗓 Année', 'callback_data': f'ods_dashboard:{view}:year'},
        ],
        [
            {'text': '📈 Graphique', 'callback_data': f'ods_dashboard_chart:{view}:{period}'},
            {'text': alternate_label, 'callback_data': f'ods_dashboard:{alternate_view}:{period}'},
        ],
    ]


def _load_workbook():
    config = legacy.microsoft_email_config()
    token = legacy.graph_access_token()
    return openpyxl.load_workbook(io.BytesIO(legacy.download_onedrive_path(
        token, config['EMAIL_SENDER'], legacy.ODS_LIST_PATH,
    )), data_only=True)


def month_picker_buttons(months, view='summary', page=0, page_size=12):
    total_pages = max(1, (len(months) + page_size - 1) // page_size)
    page = max(0, min(page, total_pages - 1))
    selected = months[page * page_size:(page + 1) * page_size]
    rows = []
    for index in range(0, len(selected), 2):
        row = []
        for key in selected[index:index + 2]:
            year, month = (int(part) for part in key.split('-'))
            row.append({
                'text': f'{MONTH_NAMES_FR[month]} {year}',
                'callback_data': f'ods_dashboard:{view}:{key}',
            })
        rows.append(row)
    navigation = []
    if page > 0:
        navigation.append({'text': '⬅️ Plus récent', 'callback_data': f'ods_dashboard_months:{view}:{page - 1}'})
    if page + 1 < total_pages:
        navigation.append({'text': 'Plus ancien ➡️', 'callback_data': f'ods_dashboard_months:{view}:{page + 1}'})
    if navigation:
        rows.append(navigation)
    rows.append([{'text': '↩️ Tableau de bord', 'callback_data': f'ods_dashboard:{view}:month'}])
    return rows, page, total_pages


def show_month_picker(chat_id, marketing=False, page=0):
    try:
        workbook = _load_workbook()
        months = available_months(workbook)
        if not months:
            legacy.tg(chat_id, 'Aucun mois disponible dans List.xlsx.')
            return
        view = 'marketing' if marketing else 'summary'
        buttons, page, total_pages = month_picker_buttons(months, view, page)
        legacy.tg(
            chat_id,
            f'🗓 Choisissez un mois — page {page + 1}/{total_pages}',
            buttons,
        )
    except Exception as exc:
        logger.exception('ODS month picker failed')
        legacy.tg(chat_id, f'❌ Liste des mois indisponible : {exc}')


def dashboard_chart(metrics):
    """Build a Telegram-ready PNG chart for the selected dashboard period."""
    plt.style.use('seaborn-v0_8-whitegrid')
    figure, axes = plt.subplots(1, 2, figsize=(11, 5.5))
    figure.patch.set_facecolor('#f7faf8')
    title = 'Marketing' if metrics['marketing'] else 'Tableau de bord'
    figure.suptitle(f"{title} — {metrics['period_label']}", fontsize=17, fontweight='bold')

    if metrics['marketing'] and metrics['sources']:
        source_rows = metrics['sources'][:8]
        names = [name for name, _ in source_rows][::-1]
        offers = [item['offers'] for _, item in source_rows][::-1]
        accepted = [item['accepted'] for _, item in source_rows][::-1]
        revenue = [item['revenue'] for _, item in source_rows][::-1]
        positions = range(len(names))
        axes[0].barh([p + 0.2 for p in positions], offers, height=0.38, label='Offres', color='#4682b4')
        axes[0].barh([p - 0.2 for p in positions], accepted, height=0.38, label='Acceptées', color='#2e8b57')
        axes[0].set_yticks(list(positions), names)
        axes[0].set_title('Performance par source')
        axes[0].legend()
        axes[1].barh(list(positions), revenue, color='#d79b32')
        axes[1].set_yticks(list(positions), names)
        axes[1].set_title('Revenus acceptés par source')
        axes[1].set_xlabel('$')
    else:
        labels = ['Acceptées', 'En cours', 'En attente', 'Refusées', 'Fermées']
        keys = ['Accept', 'In process', 'Hold', 'Refused', 'Closed']
        colors = ['#2e8b57', '#4682b4', '#d79b32', '#c84c4c', '#777777']
        counts = [metrics['statuses'].get(key, 0) for key in keys]
        axes[0].bar(labels, counts, color=colors)
        axes[0].set_title('Nombre d’offres par statut')
        axes[0].tick_params(axis='x', rotation=25)
        axes[0].set_ylabel('Nombre')
        axes[1].bar(['Soumissions', 'Acceptées'], [metrics['quoted'], metrics['revenue']], color=['#4682b4', '#2e8b57'])
        axes[1].set_title('Valeur financière')
        axes[1].set_ylabel('$')

    for axis in axes:
        axis.spines[['top', 'right']].set_visible(False)
    figure.tight_layout(rect=(0, 0, 1, 0.92))
    output = io.BytesIO()
    figure.savefig(output, format='png', dpi=160, bbox_inches='tight')
    plt.close(figure)
    output.seek(0)
    return output


def show_dashboard_chart(chat_id, period='month', marketing=False):
    try:
        metrics = dashboard_metrics(_load_workbook(), period, marketing)
        image = dashboard_chart(metrics)
        response = requests.post(
            f'https://api.telegram.org/bot{legacy.BOT_TOKEN}/sendPhoto',
            data={'chat_id': chat_id, 'caption': f"📈 {metrics['period_label']}"},
            files={'photo': ('tableau-de-bord.png', image.getvalue(), 'image/png')},
            timeout=60,
        )
        response.raise_for_status()
    except Exception as exc:
        logger.exception('ODS dashboard chart failed')
        legacy.tg(chat_id, f'❌ Graphique indisponible : {exc}')


def show_dashboard(chat_id, period='month', marketing=False):
    try:
        workbook = _load_workbook()
        metrics = dashboard_metrics(workbook, period, marketing)
        legacy.tg(chat_id, format_dashboard(metrics), dashboard_buttons(marketing, period))
    except Exception as exc:
        logger.exception('ODS dashboard generation failed')
        legacy.tg(chat_id, f"❌ Tableau de bord indisponible : {exc}")


_previous_handle_update = legacy.handle_update


def _answer_callback(cb):
    try:
        requests.post(
            f'https://api.telegram.org/bot{legacy.BOT_TOKEN}/answerCallbackQuery',
            json={'callback_query_id': cb.get('id')}, timeout=3,
        )
    except Exception:
        pass


def handle_update_dashboard(data):
    msg = data.get('message') or {}
    cb = data.get('callback_query') or {}
    actor_id = (cb.get('from') or msg.get('from') or {}).get('id')
    chat_id = ((cb.get('message') or {}).get('chat') or {}).get('id') if cb else (msg.get('chat') or {}).get('id')
    if msg and str(msg.get('text') or '').strip() == '📊 Tableau de bord':
        if actor_id not in legacy.ALLOWED_USERS:
            legacy.tg(chat_id, '⛔ Ce bot est privé.')
            return
        legacy.executor.submit(show_dashboard, chat_id, 'month', False)
        return
    cdata = str(cb.get('data') or '')
    if cdata.startswith('ods_dashboard_months:'):
        if actor_id not in legacy.ALLOWED_USERS:
            legacy.tg(chat_id, '⛔ Ce bot est privé.')
            return
        _answer_callback(cb)
        _, view, page = cdata.split(':', 2)
        try:
            page = int(page)
        except ValueError:
            page = 0
        legacy.executor.submit(show_month_picker, chat_id, view == 'marketing', page)
        return
    if cdata.startswith('ods_dashboard_chart:'):
        if actor_id not in legacy.ALLOWED_USERS:
            legacy.tg(chat_id, '⛔ Ce bot est privé.')
            return
        _answer_callback(cb)
        _, view, period = cdata.split(':', 2)
        legacy.executor.submit(show_dashboard_chart, chat_id, period, view == 'marketing')
        return
    if cdata.startswith('ods_dashboard:'):
        if actor_id not in legacy.ALLOWED_USERS:
            legacy.tg(chat_id, '⛔ Ce bot est privé.')
            return
        _answer_callback(cb)
        _, view, period = cdata.split(':', 2)
        legacy.executor.submit(show_dashboard, chat_id, period, view == 'marketing')
        return
    return _previous_handle_update(data)


legacy.handle_update = handle_update_dashboard
logger.info('ODS MANAGEMENT AND MARKETING DASHBOARD ACTIVE')
