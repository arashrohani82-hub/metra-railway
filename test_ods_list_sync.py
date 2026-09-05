import os
import io
import tempfile
from datetime import datetime

import openpyxl

os.environ.setdefault("ANTHROPIC_API_KEY", "test-key")
os.environ.setdefault("DATA_DIR", tempfile.mkdtemp(prefix="metra-ods-list-test-"))

import app


HEADERS = [
    "No", "Year ", "Month", "Description", "Price($)", "Date", "Status",
    "Accepted Price", "Source", "Contact", "Date of acceptation", "Email",
]


def workbook_with_ods_sheet():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "data 2026"
    sheet.append(HEADERS)
    sheet.append([
        1, 2026, "August", "ODS26-001-AAA-Existing", 1000,
        datetime(2026, 8, 1), "In process", 0, "Connection", "Existing", None,
        "existing@example.com",
    ])
    return workbook


def test_offer_is_inserted_once_using_full_ods_reference():
    workbook = workbook_with_ods_sheet()
    data = {
        "odsNum": "ODS26-098-RES-Correction-de-plans",
        "price": 2500,
        "name": "Mme Christine Vaitekunas",
        "email": "client@example.com",
        "email_sent_at": "2026-08-19T14:00:00",
    }

    first_row = app.upsert_ods_list_workbook(workbook, data, "In process")
    second_row = app.upsert_ods_list_workbook(workbook, data, "In process")

    assert first_row == second_row == 3
    assert workbook["data 2026"].max_row == 3
    assert workbook["data 2026"]["D3"].value.startswith("ODS26-098-RES")
    assert workbook["data 2026"]["G3"].value == "In process"


def test_project_conversion_marks_offer_accepted():
    workbook = workbook_with_ods_sheet()
    data = {
        "odsNum": "ODS26-099-AGR-Agrandissement",
        "price": 2800,
        "name": "Veronique De Foy",
        "email": "ceo@example.com",
        "email_sent_at": "2026-08-19T14:00:00",
    }
    app.upsert_ods_list_workbook(workbook, data, "In process")
    accepted_at = datetime(2026, 8, 28, 10, 30)

    row = app.upsert_ods_list_workbook(workbook, data, "Accept", accepted_at)
    sheet = workbook["data 2026"]

    assert row == 3
    assert sheet["G3"].value == "Accept"
    assert sheet["H3"].value == 2800
    assert sheet["K3"].value == accepted_at


def test_project_conversion_records_project_folder_for_guardian_sync():
    workbook = workbook_with_ods_sheet()
    data = {
        "odsNum": "ODS26-100-STR-Plans-structuraux",
        "project_folder": "P26-031-STR-Plans-structuraux",
        "price": 4200,
        "name": "Client Test",
    }
    row = app.upsert_ods_list_workbook(workbook, data, "Accept")
    sheet = workbook["data 2026"]
    headers = {str(cell.value or "").strip(): cell.column for cell in sheet[1]}

    assert sheet.cell(row, headers["Project Code"]).value == data["project_folder"]


def test_missing_calculation_properties_are_repaired_before_save():
    workbook = workbook_with_ods_sheet()
    workbook.calculation = None
    data = {
        "odsNum": "ODS26-107-GGH-Évaluation-structurale-des-fondations",
        "price": 1800,
        "name": "Nevin El-Tahry",
        "email": "nevin.i.reda@gmail.com",
        "email_sent_at": "2026-08-31T11:51:00",
    }

    row = app.upsert_ods_list_workbook(workbook, data, "In process")
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    reopened = openpyxl.load_workbook(output)

    assert row == 3
    assert reopened.calculation is not None
    assert reopened.calculation.fullCalcOnLoad is True
    assert reopened.calculation.forceFullCalc is True
    assert reopened.calculation.calcMode == "auto"
    assert reopened["data 2026"]["D3"].value.startswith("ODS26-107-GGH")
