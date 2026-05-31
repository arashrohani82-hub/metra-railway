import os, json, io, shutil, logging
from datetime import datetime
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, ContextTypes, filters
import anthropic
import openpyxl

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

ANTHROPIC_KEY = os.environ.get('ANTHROPIC_API_KEY')
BOT_TOKEN = os.environ.get('TELEGRAM_BOT_TOKEN')
TEMPLATE = '/app/template.xlsx'
BASE = os.path.dirname(os.path.abspath(__file__))

client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)

# Price list by service
PRICES = {
    'Analyse structurale générale': 3500,
    'Inspection et rapport structural': 2800,
    "Avis d'expert — stabilisation et renforcement": 3200,
    'Enlèvement de mur porteur': 3800,
    'Inspection des fondations': 2500,
    'Évaluation des fissures et désordres structuraux': 2200,
    'Mur de soutènement': 4200,
    'Conception structurale complète': 6500,
    'Analyse structurale — sous-sol et ajout au-dessus du garage': 4500,
    'Réaménagement intérieur avec modification structurale': 3500,
}

user_data = {}

async def start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 Bienvenue — Métra Structure ODS Generator\n\n"
        "📸 Envoyez une photo du client (SoumissionRenovation, courriel, formulaire)\n"
        "Je vais extraire les informations et générer l'ODS automatiquement."
    )

async def handle_photo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    await update.message.reply_text("🔍 Extraction des informations en cours...")

    # Download photo
    photo = update.message.photo[-1]
    f = await photo.get_file()
    img_bytes = await f.download_as_bytearray()
    import base64
    img_b64 = base64.b64encode(img_bytes).decode()

    prompt = """This is a client document (SoumissionRenovation.ca screenshot, email, or form).
Extract all client and project information. Return ONLY a valid JSON object:
{"client_name":"","phone":"","email":"","address":"","soumission_ref":"","project_description":"","property_type":"","suggested_service":"","suggested_price":0}
For suggested_service choose from: "Analyse structurale générale","Inspection et rapport structural","Avis d'expert — stabilisation et renforcement","Enlèvement de mur porteur","Inspection des fondations","Évaluation des fissures et désordres structuraux","Mur de soutènement","Conception structurale complète","Analyse structurale — sous-sol et ajout au-dessus du garage","Réaménagement intérieur avec modification structurale".
suggested_price: realistic CAD integer. ONLY JSON."""

    try:
        response = client.messages.create(
            model="claude-sonnet-4-6", max_tokens=1000,
            messages=[{"role":"user","content":[
                {"type":"image","source":{"type":"base64","media_type":"image/jpeg","data":img_b64}},
                {"type":"text","text":prompt}
            ]}]
        )
        text = ''.join(b.text for b in response.content if hasattr(b,'text'))
        info = json.loads(text.replace('```json','').replace('```','').strip())

        # Generate ODS number
        yr = datetime.now().strftime('%y')
        import random
        ods_num = f"ODS{yr}-{random.randint(100,999)}"

        # Store data
        price = info.get('suggested_price') or PRICES.get(info.get('suggested_service',''), 3200)
        user_data[uid] = {
            'name': info.get('client_name',''),
            'phone': info.get('phone',''),
            'email': info.get('email',''),
            'addr': info.get('address',''),
            'ref': info.get('soumission_ref',''),
            'desc': info.get('project_description',''),
            'type': info.get('property_type',''),
            'service': info.get('suggested_service','Analyse structurale générale'),
            'price': price,
            'ods_num': ods_num,
        }

        msg = (
            f"✅ *Informations extraites*\n\n"
            f"👤 *Client:* {info.get('client_name','—')}\n"
            f"📍 *Adresse:* {info.get('address','—')}\n"
            f"📞 *Tél:* {info.get('phone','—')}\n"
            f"📧 *Courriel:* {info.get('email','—')}\n"
            f"🏠 *Type:* {info.get('property_type','—')}\n"
            f"🔢 *Réf:* {info.get('soumission_ref','—')}\n\n"
            f"🔧 *Service:* {info.get('suggested_service','—')}\n"
            f"💰 *Prix suggéré:* ${price:,} CAD\n"
            f"📄 *N° ODS:* {ods_num}\n\n"
            f"Confirmez-vous ces informations?"
        )

        keyboard = [
            [InlineKeyboardButton("✅ Confirmer & générer Excel", callback_data="confirm")],
            [InlineKeyboardButton("✏️ Modifier le prix", callback_data="change_price")],
        ]
        await update.message.reply_text(msg, parse_mode='Markdown',
                                        reply_markup=InlineKeyboardMarkup(keyboard))
    except Exception as e:
        await update.message.reply_text(f"❌ Erreur: {str(e)}")

async def handle_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    uid = query.from_user.id
    await query.answer()

    if query.data == "confirm":
        await query.edit_message_text("⏳ Génération du fichier Excel...")
        await generate_excel(query, uid)

    elif query.data == "change_price":
        await query.edit_message_text(
            "💰 Entrez le nouveau prix (nombre seulement, ex: 3500):"
        )
        ctx.user_data['waiting_price'] = True

async def handle_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if ctx.user_data.get('waiting_price') and uid in user_data:
        try:
            price = int(update.message.text.strip().replace('$','').replace(',','').replace(' ',''))
            user_data[uid]['price'] = price
            ctx.user_data['waiting_price'] = False

            keyboard = [[InlineKeyboardButton("✅ Confirmer & générer Excel", callback_data="confirm")]]
            await update.message.reply_text(
                f"💰 Prix mis à jour: ${price:,} CAD\n\nConfirmez?",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
        except:
            await update.message.reply_text("❌ Entrez un nombre valide (ex: 3500)")
    else:
        await update.message.reply_text("📸 Envoyez une photo du client pour commencer.")

async def generate_excel(query, uid):
    d = user_data.get(uid)
    if not d:
        await query.message.reply_text("❌ Données introuvables.")
        return
    try:
        template = os.path.join(BASE, 'template.xlsx')
        output = f"/tmp/ODS_{d['ods_num']}.xlsx"
        shutil.copy(template, output)

        wb = openpyxl.load_workbook(output)
        ws = wb['ODS']

        ws['B7'] = f"M./Mme {d['name']}"
        ws['B8'] = f"Adresse: :{d['addr']}"
        ws['B9'] = f"Cell.: {d['phone']}"
        ws['B10'] = f"Courriel : {d['email']}"
        ws['B12'] = f"{d['ods_num']}-{d['name'].replace(' ','-')}-{d['service'][:30]}"
        ws['B47'] = d['desc']
        ws['C47'] = 'Forfait'
        ws['D47'] = 1
        ws['E47'] = d['price']
        ws['F47'] = '=E47*D47'
        ws['F48'] = '=SUM(F47:F47)'

        wb.save(output)

        filename = f"{d['ods_num']}_{d['name'].replace(' ','-')}.xlsx"
        with open(output, 'rb') as f:
            await query.message.reply_document(
                document=f,
                filename=filename,
                caption=f"✅ ODS généré!\n📄 {filename}\n💰 ${d['price']:,} CAD"
            )
        os.remove(output)
    except Exception as e:
        await query.message.reply_text(f"❌ Erreur génération: {str(e)}")

def main():
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    app.add_handler(CallbackQueryHandler(handle_callback))
    logger.info("Bot started...")
    app.run_polling(drop_pending_updates=True)

if __name__ == '__main__':
    main()
